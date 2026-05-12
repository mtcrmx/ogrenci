"""
export.py
---------
Excel rapor üretim modülü.
Yalnızca öğrenci numarası ve emoji içermeyen ihlal özeti.
Gerekli: pip install openpyxl
"""

from __future__ import annotations
import re
from collections import Counter

from database import KRITERLER, sinif_ogrencileri, ogrenci_tik_gecmisi

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


RENK = {
    "sutun_bg":    "FF2563EB",
    "sutun_yaz":   "FFFFFFFF",
    "alt_satir":   "FFF0F4FF",
    "siyah":       "FF1E293B",
}


def _dolu_dolgu(argb: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=argb)


def _ince_kenar() -> Border:
    ince = Side(style="thin", color="FFCBD5E1")
    return Border(left=ince, right=ince, top=ince, bottom=ince)


def _sutun_gen(ws, col: int, gen: float):
    ws.column_dimensions[get_column_letter(col)].width = gen


def _kriter_saf_metin(kriter: str) -> str:
    """Tik kaydındaki metinden emoji ve gereksiz önekleri kaldırır; yalnızca ihlal açıklaması kalır."""
    s = (kriter or "").strip()
    if not s:
        return ""
    for em, ad in KRITERLER:
        if s.startswith(f"{em} ") or s.startswith(f"{em}\u200d ") or s.startswith(f"{em}\uFE0F "):
            return ad.strip()
        if s == ad:
            return ad.strip()
    # Bilinen önek yoksa baştaki emoji bloklarını sil
    out = re.sub(
        r"[\U0001F300-\U0001FAFF\U00002600-\U000027BF\U0001F1E6-\U0001F1FF"
        r"\U0000FE0F\U0000200D]+",
        "",
        s,
    )
    out = re.sub(r"\s+", " ", out).strip(" ,;")
    return out if out else s


def _ihlal_ozeti_yazisi(gecmis: list[dict]) -> str:
    """Her ihlal türü için temiz metin + tekrar sayısı; emoji yok."""
    say = Counter()
    for k in gecmis:
        temiz = _kriter_saf_metin(k.get("kriter") or "")
        if temiz:
            say[temiz] += 1
    if not say:
        return "İhlal kaydı yok"
    parcalar = [
        f"{ad} ({n})"
        for ad, n in sorted(say.items(), key=lambda x: (-x[1], x[0]))
    ]
    return ", ".join(parcalar)


def excel_raporu_olustur(
    kayit_yolu: str,
    ogretmen_adi: str,
    sinif_listesi: list[dict],
    yalnizca_sinif_id: int | None = None,
) -> str:
    """
    Tek sayfa: Öğrenci No | İhlal Özeti (yalnızca metin, emoji yok).
    """
    if not OPENPYXL_OK:
        raise ImportError(
            "openpyxl kütüphanesi bulunamadı.\n"
            "Lütfen terminalde şunu çalıştırın:\n  pip install openpyxl"
        )

    if yalnizca_sinif_id is not None:
        hedef_siniflar = [s for s in sinif_listesi if s["id"] == yalnizca_sinif_id]
    else:
        hedef_siniflar = sinif_listesi

    tum: list[dict] = []
    for sinif in hedef_siniflar:
        for ogr in sinif_ogrencileri(sinif["id"]):
            ogr = dict(ogr)
            ogr["sinif_adi"] = sinif["sinif_adi"]
            tum.append(ogr)

    tum.sort(key=lambda o: (o.get("sinif_adi") or "", int(o.get("ogr_no") or 0)))

    wb = Workbook()
    ws = wb.active
    ws.title = "Ihlal Ozeti"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    basliklar = ["Öğrenci No", "İhlal özeti"]
    for col, b in enumerate(basliklar, 1):
        cell = ws.cell(row=1, column=col, value=b)
        cell.font = Font(bold=True, size=11, color=RENK["sutun_yaz"])
        cell.fill = _dolu_dolgu(RENK["sutun_bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _ince_kenar()
    ws.row_dimensions[1].height = 24

    _sutun_gen(ws, 1, 14)
    _sutun_gen(ws, 2, 72)

    satir = 2
    for idx, ogr in enumerate(tum):
        gecmis = ogrenci_tik_gecmisi(ogr["id"])
        ozet = _ihlal_ozeti_yazisi(gecmis)
        no = ogr.get("ogr_no")
        alt_bg = RENK["alt_satir"] if idx % 2 == 0 else "FFFFFFFF"

        c1 = ws.cell(row=satir, column=1, value=no)
        c1.fill = _dolu_dolgu(alt_bg)
        c1.border = _ince_kenar()
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c1.font = Font(size=11, color=RENK["siyah"])

        c2 = ws.cell(row=satir, column=2, value=ozet)
        c2.fill = _dolu_dolgu(alt_bg)
        c2.border = _ince_kenar()
        c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c2.font = Font(size=10, color=RENK["siyah"])

        ws.row_dimensions[satir].height = max(18, min(60, 14 + ozet.count(",") * 4))
        satir += 1

    if satir > 2:
        ws.auto_filter.ref = f"A1:B{satir - 1}"

    wb.save(kayit_yolu)
    return kayit_yolu
