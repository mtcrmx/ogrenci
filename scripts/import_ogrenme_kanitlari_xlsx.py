# -*- coding: utf-8 -*-
"""Excel'deki TYMM öğrenme kanıtları tablolarını ödev takibi JSON'una dönüştürür."""
from __future__ import annotations

import argparse
import json
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

SHEET_TO_DERS = {
    "türkçe": "Türkçe",
    "turkce": "Türkçe",
    "mat": "Matematik",
    "fen": "Fen Bilimleri",
    "sos": "Sosyal Bilgiler",
    "din": "Din Kültürü",
    "ing": "İngilizce",
    "i̇ng": "İngilizce",
}

TEMA_RE = re.compile(
    r"^\s*(\d+)\.\s*TEMA\s*:\s*(.+)$",
    re.IGNORECASE | re.UNICODE,
)
ING_UNITE_THEME_RE = re.compile(
    r"^\s*[UuÜü]nite\s*\d+\s*[–—\-]\s*Theme\s*:\s*.+",
    re.UNICODE,
)
THEME_EN_RE = re.compile(
    r"^\s*Theme\s*(\d+)\s*:\s*(.+)$",
    re.IGNORECASE | re.UNICODE,
)


def _norm_sheet_name(name: str) -> str:
    if not name:
        return ""
    n = name.strip().lower()
    n = n.replace("ı", "i").replace("İ", "i")
    return n


def _map_ders(sheet_name: str) -> str | None:
    key = _norm_sheet_name(sheet_name)
    for prefix, ders in SHEET_TO_DERS.items():
        if key == prefix or key.startswith(prefix):
            return ders
    # Türkçe sayfa adı Tam genişlik harflerle olabilir
    aliases = {
        "trke": "Türkçe",
        "matematik": "Matematik",
        "fen bilimleri": "Fen Bilimleri",
        "sosyal": "Sosyal Bilgiler",
        "din": "Din Kültürü",
        "ingilizce": "İngilizce",
    }
    return aliases.get(key)


def _ascii_fold(s: str) -> str:
    """Türkçe karakterleri kabaca ASCII'ye indirger (tema / ünite eşlemesi için)."""
    t = unicodedata.normalize("NFD", (s or "").casefold())
    t = "".join(ch for ch in t if unicodedata.category(ch) != "Mn")
    for a, b in (
        ("ı", "i"),
        ("ğ", "g"),
        ("ü", "u"),
        ("ş", "s"),
        ("ö", "o"),
        ("ç", "c"),
        ("â", "a"),
        ("î", "i"),
        ("û", "u"),
    ):
        t = t.replace(a, b)
    return t


def _is_tema_line(s: str) -> bool:
    """TEMA, Öğrenme alanı, Ünite veya İngilizce Theme satırları."""
    s = (s or "").strip()
    if not s:
        return False
    if TEMA_RE.match(s) or THEME_EN_RE.match(s) or ING_UNITE_THEME_RE.match(s):
        return True
    m = re.match(r"^\s*\d+\.\s*(.+?)\s*:\s*.+", s)
    if not m:
        return False
    head = _ascii_fold(m.group(1))
    compact = head.replace(" ", "")
    if "tema" in compact:
        return True
    if "ogrenmealani" in compact:
        return True
    if "unite" in compact:
        return True
    if head.strip().startswith("theme"):
        return True
    return False


def _parse_tema_title(s: str) -> str:
    s = (s or "").strip()
    m = ING_UNITE_THEME_RE.match(s)
    if m:
        inner = s.strip()
        return inner
    m = TEMA_RE.match(s)
    if m:
        return f"{m.group(1)}. Tema: {m.group(2).strip()}"
    m = THEME_EN_RE.match(s)
    if m:
        return f"Theme {m.group(1)}: {m.group(2).strip()}"
    m = re.match(r"^\s*(\d+)\.\s*(.+?)\s*:\s*(.+)$", s)
    if m:
        return f"{m.group(1)}. {m.group(2).strip()}: {m.group(3).strip()}"
    return s


def parse_sheet(
    ws,
    grades: list[int],
    konu_basligi: str = "Öğrenme kanıtları (ölçme ve değerlendirme)",
) -> dict[int, list[dict]]:
    """Her sütun bir sınıf düzeyine karşılık gelir (grades[i] = sütun i)."""
    out: dict[int, list[dict]] = {g: [] for g in grades}
    max_col = min(len(grades), ws.max_column)

    for col_idx in range(1, max_col + 1):
        grade = grades[col_idx - 1]
        current_tema: str | None = None
        current_ciktilar: list[str] = []

        def flush():
            nonlocal current_tema, current_ciktilar
            if not current_tema:
                return
            baslik = konu_basligi
            out[grade].append(
                {
                    "tema": f"{grade}. sınıf · {_parse_tema_title(current_tema)}",
                    "konular": [{"baslik": baslik, "ciktilar": current_ciktilar}],
                }
            )
            current_ciktilar = []

        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx).value
            if cell is None:
                continue
            text = str(cell).strip()
            if not text:
                continue
            # Başlık satırları (yüzde / adet bilgisi)
            if "Öğrenme Kanıtları" in text or "Sosyal-Duygusal" in text:
                continue
            # Sadece sınıf düzeyi başlığı (ör. "5. Sınıf", "5.Sınıf")
            if "sinif" in _ascii_fold(text).replace(" ", "") and re.match(
                r"^\s*\d+\.", text
            ):
                continue
            ac = _ascii_fold(text).replace(" ", "").replace(":", "")
            if ac.startswith("ogrenmekanitlari") and len(ac) < 35:
                continue
            if _is_tema_line(text):
                flush()
                current_tema = text
                current_ciktilar = []
                continue
            if current_tema:
                # Aynı hücrede tema + kanıt karışık değil; tekrarları atla
                if text not in current_ciktilar:
                    current_ciktilar.append(text)
        flush()

    return out


def build_ders_tree(
    all_sheets: dict[str, dict[int, list[dict]]],
) -> dict[str, list[dict]]:
    """sinif_seviyesi -> tema listesi JSON için: ders başına 5-8 birleşik sıra."""
    result: dict[str, list[dict]] = {}
    for ders, by_grade in all_sheets.items():
        merged: list[dict] = []
        for g in (5, 6, 7, 8):
            merged.extend(by_grade.get(g, []))
        if merged:
            result[ders] = merged
    return result


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", type=Path, help="Excel dosyası")
    ap.add_argument(
        "-o",
        "--output",
        type=Path,
        default=Path(__file__).resolve().parent.parent / "data" / "ogrenme_kanitlari_excel.json",
    )
    args = ap.parse_args()

    wb = load_workbook(args.xlsx, data_only=True)
    all_sheets: dict[str, dict[int, list[dict]]] = {}
    grades = [5, 6, 7, 8]

    for sn in wb.sheetnames:
        ders = _map_ders(sn)
        if not ders:
            continue
        konu_b = (
            "Öğrenme çıktıları"
            if ders == "İngilizce"
            else "Öğrenme kanıtları (ölçme ve değerlendirme)"
        )
        parsed = parse_sheet(wb[sn], grades, konu_basligi=konu_b)
        all_sheets[ders] = parsed

    doc = {
        "kaynak": str(args.xlsx),
        "aciklama": "Excel TYMM öğrenme kanıtları (ölçme ve değerlendirme) — ödev çıktı listeleri için.",
        "dersler": build_ders_tree(all_sheets),
    }

    args.output.parent.mkdir(parents=True, exist_ok=True)
    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(doc, f, ensure_ascii=False, indent=2)

    print(f"Yazıldı: {args.output} ({len(doc['dersler'])} ders)")


if __name__ == "__main__":
    main()
