"""
web_app.py  —  Erenler Cumhuriyet Ortaokulu Ogrenci Takip
(Flask rotalari <int:...> ile tam; GitHub/Render senkron)
"""

import json
import os, tempfile
from datetime import datetime
from io import BytesIO
from flask import (
    Flask, render_template, request, redirect, url_for,
    session, jsonify, send_file, make_response, flash, abort,
)
from database import (
    ogrenci_ozellikleri_getir, ogrenci_ozellik_artir,
    initialize_db, KRITERLER, OLUMLU_KRITERLER,
    tum_ogretmenler, tum_sifre_listesi,
    ogretmen_dogrula, ogretmen_id_bul, ogretmen_siniflari,
    sinif_ogrencileri, tum_okul_ogrencileri, ogrenci_tik_gecmisi,
    ogrenci_tik_sayisi, OLUMSUZ_TIK_LIMIT, OLUMLU_TIK_LIMIT,
    tik_ekle, tek_ogrenci_sifirla, sinif_sifirla, tum_tikleri_sifirla,
    ogretmenin_ogrenci_tiklerini_sifirla, ogretmenin_sinif_tiklerini_sifirla,
    olumlu_tik_ekle, olumlu_sinif_etkinlik_ekle,
    ogrenci_olumlu_tik_sayisi, ogrenci_olumlu_tik_sayilari,
    lig_siralama, lig_manuel_sifirla, sinif_olumlu_gecmis,
    gunluk_mac_olustur, bugun_maclar, lig_puan_tablosu, lig_mac_tablo_sifirla,
    mac_sonucu_gir, mac_oy_ver, kart_ver, mac_kartlari,
    # Gamification
    SEVIYELER, ROZET_TANIMI, SANS_CARKI_SEENEKLERI,
    MUFETTIS_YETKILILERI,
    sinif_seviye_hesapla, tum_sinif_seviyeleri,
    rozet_ver_ogrenci, son_rozetler,
    sinif_seri_guncelle, tum_seri_tablosu,
    bugun_gorev, gorev_tamamla,
    bugun_mufettis, mufettis_degerlendir, mufettis_belirle,
    alkis_ver, son_alkislar,
    sezon_puan_ekle, sezon_siralama,
    ittifak_olustur, aktif_ittifaklar, ittifak_tamamla,
    ittifak_ogrenci_talebi, ittifak_onayla, ittifak_reddet,
    bekleyen_ogrenci_talepleri,
    tum_verileri_sifirla, sistem_yedek_listesi, sistem_yedegini_geri_yukle,
    quiz_sorular_getir, quiz_sonuc_kaydet, quiz_gunluk_dersleri,
    quiz_sinif_istatistik, quiz_sorulari_yukle,
    tik_dondur,
    odev_ekle, sinif_odevleri, odev_detay as odev_detay_db, odev_tamamla,
    odev_tamamlandi_kaldir, ogrenci_odevleri,
    gelisim_ozeti, gelisim_gorev_tamamla, sandik_ac, telafi_gorevi_olustur,
    tebrik_gonder, haftalik_veli_ozeti,
    akilli_ogrenci_karnesi, ogretmen_bildirim_merkezi, gelisim_ligi,
    hikaye_modu, pazar_urunleri_ogrenci, pazar_satin_al, ogretmen_notu_ekle,
    ogrenci_rozetleri_yayin_map, rozet_emojileri_ve_metin,
    envanter_listele, envanter_aktif_ayarla, ogrenci_aktif_envanter_map,
    oyun_puani_kaydet, GOREV_SABLONLARI,
    MEVKILER, taktik_yukle, taktik_kaydet, spor_taktik_yukle, spor_taktik_kaydet,
    ogrenci_mac_olustur, ogrenci_mac_listesi, ogrenci_mac_detay,
    ogretmen_onay_bekleyen_ogrenci_maclari, ogrenci_mac_onayla,
    bilgilendirme_ekle, bilgilendirme_listesi, bilgilendirme_yayinlayan_icin_sil,
    son_bilgilendirme,
    rapor_arsiv_kaydet, rapor_arsiv_listesi, rapor_arsiv_pdf_oku,
    rapor_arsiv_tumunu_yedekle_ve_sil, rapor_arsiv_yedek_gruplari, rapor_arsiv_grubu_geri_yukle,
    tik_kayitlari_siniflarda,
    ogretmen_yetki_al, ogretmen_yetki_guncelle,
    randevu_talep_ekle, randevu_talep_by_id, randevu_listesi_siniflar, randevu_durum_guncelle,
    gunluk_yansima_ekle, gunluk_yansima_by_id, gunluk_yansima_bekleyen_siniflar,
    gunluk_yansima_degerlendir,
    gunluk_yansima_ogrenci_gecmis,
    haftalik_sinif_ozeti, tik_sayisi_sinif_aralik, olumlu_sayisi_sinif_aralik,
    veli_ozet_metrikleri, anonim_sinif_dagilimi,
    denetim_listesi, denetim_kaydet, admin_meta_get, admin_meta_set,
)
from export import excel_raporu_olustur, OPENPYXL_OK
from pdf_export import PDF_OK, derle_analiz_snapshot, pdf_analiz_uret_bytes
from rapor_analiz import (
    aylik_tik_sayilari,
    durum_dagilimi_hesapla,
    kriter_dagilimi_satirlardan,
    oneriler_derle,
    sinif_en_yuksek_ortalama_bul,
)

_BASE = os.path.dirname(os.path.abspath(__file__))
app = Flask(__name__,
            template_folder=os.path.join(_BASE, "templates"),
            static_folder=os.path.join(_BASE, "static"))
app.secret_key = os.environ.get("SECRET_KEY", "erenler-cumhuriyet-2025-gizli")

SIFIR_PAROLA = "1234"
ADMIN_SIFRE  = "ECadmin"
OGRENCI_SIFRE = os.environ.get("OGRENCI_SIFRE", "ogrenci")
VELI_SIFRE = os.environ.get("VELI_SIFRE", "veli")

initialize_db()


def _tarih_araligi_duzelt(bas: str, bit: str) -> tuple[str, str]:
    """ISO yyyy-mm-dd aralığında başlangıç > bitiş ise tarihleri yer değiştirir."""
    if not bas or not bit:
        return bas, bit
    try:
        da = datetime.strptime(bas.strip()[:10], "%Y-%m-%d").date()
        db = datetime.strptime(bit.strip()[:10], "%Y-%m-%d").date()
        if da > db:
            return db.isoformat(), da.isoformat()
        return da.isoformat(), db.isoformat()
    except ValueError:
        return bas, bit


def _ogretmen_sinifinda_mi(ogretmen_id: int, sinif_id: int) -> bool:
    return sinif_id in {s["id"] for s in ogretmen_siniflari(ogretmen_id)}


def _ogretmen_ogrencisine_erisebilir(ogretmen_id: int, ogrenci_id: int) -> bool:
    og = _ogrenci_bul(ogrenci_id)
    if not og:
        return False
    return _ogretmen_sinifinda_mi(ogretmen_id, int(og["sinif_id"]))


# Okul geneli sıfırlama (tüm tikler, lig sezonu, tam veri silme) — yalnızca bu öğretmen.
_TOPLU_SIFIRLAMA_AD_SOYAD = "ADEM AKGÜL"


def _toplu_sifirlamaya_izinli_mi(ogretmen_id: int) -> bool:
    beklenen = ogretmen_id_bul(_TOPLU_SIFIRLAMA_AD_SOYAD)
    return beklenen is not None and beklenen == ogretmen_id


def _ogretmen_ogrenci_macina_erisebilir(ogretmen_id: int, mac: dict | None) -> bool:
    if not mac:
        return False
    if _toplu_sifirlamaya_izinli_mi(ogretmen_id):
        return True
    siniflar = {s["id"] for s in ogretmen_siniflari(ogretmen_id)}
    return int(mac["sinif1_id"]) in siniflar or int(mac["sinif2_id"]) in siniflar


# Yalnızca rapor/analiz görebilen öğretmenler (`yetki=rapor`) bu endpoint’lere girebilir;
# tik/ödev vb. diğerleri rapor_ozet’e yönlendirilir.
_RAPOR_SADECE_ROTALAR = frozenset({
    "rapor_ozet", "rapor_ozet_csv", "rapor_excel", "rapor_excel_detayli",
    "rapor_analiz_pdf", "rapor_arsiv_sayfa", "rapor_arsiv_sifirla",
    "rapor_arsiv_yedek_geri_yukle", "rapor_arsiv_indir",
    "analiz_merkezi",
    "rapor_haftalik", "rapor_karsilastir", "rapor_anonim_sinif",
    "manifest", "service_worker",
})


# ── Yardimcilar ──────────────────────────────────────────────────────────────
def giris_zorunlu(fn):
    from functools import wraps
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if "ogretmen_id" not in session:
            return redirect(url_for("login"))
        if session.get("ogretmen_id") and session.get("ogretmen_yetki") is None:
            session["ogretmen_yetki"] = ogretmen_yetki_al(session["ogretmen_id"])
        if session.get("ogretmen_yetki") == "rapor":
            ep = request.endpoint
            if ep and ep not in _RAPOR_SADECE_ROTALAR:
                flash(
                    "Bu bölüm için tam öğretmen yetkisi gerekir. Size yalnızca raporlar açık.",
                    "warning",
                )
                return redirect(url_for("rapor_ozet"))
        return fn(*args, **kwargs)
    return wrapper


def ogrenci_giris_zorunlu(fn):
    from functools import wraps
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not session.get("ogrenci_giris"):
            if request.path.startswith("/api/"):
                return jsonify({"ok": False, "sebep": "Ogrenci sifresi gerekli"}), 401
            return redirect(url_for("ogrenci_giris", next=request.path))
        return fn(*args, **kwargs)
    return wrapper


def bilgilendirme_yetkili_mi() -> bool:
    return "ogretmen_id" in session


def _bilgilendirme_hedefi() -> str | None:
    if session.get("veli_ogrenci_id"):
        return "veli"
    if session.get("ogrenci_giris"):
        return "ogrenci"
    if session.get("ogretmen_id"):
        return "ogretmen"
    return None


def _bilgilendirme_ana_ekran_mi(hedef: str) -> bool:
    ana_ekranlar = {
        "ogretmen": {"dashboard"},
        "ogrenci": {"ogrenci_gorunum", "ogrenci_gelisim_panel"},
        "veli": {"veli_panel"},
    }
    return request.endpoint in ana_ekranlar.get(hedef, set())


def _alan_tanitimi_verisi(hedef: str) -> dict | None:
    veriler = {
        "ogretmen": {
            "baslik": "Öğretmen paneline hoş geldiniz",
            "alt": "Bu sürümde yalnızca davranış tik kaydı kullanılır.",
            "adimlar": [
                {"etiket": "Ana Menü", "metin": "Sol listeden sınıf seçin; öğrenci kartından tik atın veya geçmişe bakın."},
                {"etiket": "Tik geçmişi", "metin": "Karttaki 📜 ile öğrencinin tik kayıtlarını görürsünüz."},
                {"etiket": "Excel", "metin": "Üst menüden sınıf listesini tik durumuyla dışa aktarabilirsiniz."},
            ],
        },
        "ogrenci": {
            "baslik": "Öğrenci alanına hoş geldin",
            "alt": "Yalnızca kendi tik kayıtlarını ve nedenlerini görüntüleyebilirsin.",
            "adimlar": [
                {"etiket": "Tik listesi", "metin": "Her satırda tik tarihi, öğretmenin seçtiği davranış nedeni (kriter) ve kaydı işleyen öğretmen adı yer alır."},
            ],
        },
        "veli": {
            "baslik": "Veli ekranına hoş geldiniz",
            "alt": "Yalnızca öğrencinizin tik kayıtlarını ve nedenlerini görüntüleyebilirsiniz.",
            "adimlar": [
                {"etiket": "Tik geçmişi", "metin": "Her satırda tik tarihi, öğretmenin seçtiği davranış nedeni (kriter) ve kaydı işleyen öğretmen adı yer alır."},
            ],
        },
    }
    return veriler.get(hedef)


def _odev_bildirimi_verisi(hedef: str) -> dict | None:
    if hedef == "ogrenci":
        ogrenci_id = session.get("ogrenci_id")
        hedef_adi = "Öğrenci"
    elif hedef == "veli":
        ogrenci_id = session.get("veli_ogrenci_id")
        hedef_adi = "Veli"
    else:
        return None
    if not ogrenci_id:
        return None
    odevler = ogrenci_odevleri(int(ogrenci_id), 30)
    if not odevler:
        return None
    bekleyen = [o for o in odevler if not o.get("tamamlandi")]
    if not bekleyen:
        return None
    son_odev = max(bekleyen, key=lambda o: int(o.get("id") or 0))
    return {
        "hedef": hedef_adi,
        "odev": son_odev,
        "ogrenci_id": int(ogrenci_id),
    }


@app.context_processor
def _bilgilendirme_context():
    return {
        "bilgilendirme_yetkili": bilgilendirme_yetkili_mi(),
    }


@app.after_request
def _bilgilendirme_modal_ekle(response):
    try:
        if response.direct_passthrough:
            return response
        if not (response.content_type or "").startswith("text/html"):
            return response
        hedef = _bilgilendirme_hedefi()
        if not hedef:
            return response
        if not _bilgilendirme_ana_ekran_mi(hedef):
            return response
        parcaciklar = []
        bilgi = son_bilgilendirme(hedef)
        if bilgi:
            parcaciklar.append(render_template(
                "_bilgilendirme_modal.html",
                aktif_bilgilendirme=bilgi,
                bilgilendirme_hedef=hedef,
            ))
        tanitim = _alan_tanitimi_verisi(hedef)
        if tanitim:
            parcaciklar.append(render_template(
                "_alan_tanitimi.html",
                tanitim=tanitim,
                tanitim_hedef=hedef,
            ))
        odev_bildirimi = _odev_bildirimi_verisi(hedef)
        if odev_bildirimi:
            parcaciklar.append(render_template(
                "_odev_bildirimi_modal.html",
                odev_bildirimi=odev_bildirimi,
                odev_hedef=hedef,
            ))
        if not parcaciklar:
            return response
        html = response.get_data(as_text=True)
        modal = "\n".join(parcaciklar)
        if "</body>" in html:
            html = html.replace("</body>", modal + "\n</body>")
        else:
            html += modal
        response.set_data(html)
        response.headers["Content-Length"] = len(response.get_data())
    except Exception:
        pass
    return response


# Tik seviye sistemi: 3-Uyari / 6-Veli / 9-Tutanak / 12-Disiplin
TIK_SEVIYELERI = [
    (12, "disiplin",  "🚨", "Disiplin Cezasi"),
    (9,  "tutanak",   "📋", "Tutanak"),
    (6,  "veli",      "📱", "Veli Bilgilendirme"),
    (3,  "uyari",     "⚠️",  "Uyari"),
    (0,  "temiz",     "✅",  None),
]

def _durum(tik: int) -> dict:
    for esik, kod, emoji, etiket in TIK_SEVIYELERI:
        if tik >= esik:
            return {"kod": kod, "emoji": emoji, "etiket": etiket,
                    "basamak": esik if esik > 0 else None}
    return {"kod": "temiz", "emoji": "✅", "etiket": None, "basamak": None}


def _var_hakem_idleri() -> list[int]:
    from database import VAR_INCELEME_OGRETMENLER

    out: list[int] = []
    for ad in VAR_INCELEME_OGRETMENLER:
        oid = ogretmen_id_bul(ad)
        if oid:
            out.append(int(oid))
    return out


LIG_KART_NEDENLERI = [
    "Kural ihlali",
    "Saygısız davranış",
    "Erken uyarı / tekerrür",
    "Takım oyununa aykırı hareket",
]


def _ogrencilere_durum_ekle(liste: list[dict]) -> list[dict]:
    for o in liste:
        d = _durum(o["tik_sayisi"])
        o["durum"]   = d["kod"]
        o["emoji"]   = d["emoji"]
        o["etiket"]  = d["etiket"]
        o["basamak"] = d["basamak"]
    return liste


def _tum_siniflar() -> list[dict]:
    from database import _conn as _db
    con = _db()
    rows = [dict(r) for r in con.execute(
        "SELECT id, sinif_adi FROM siniflar ORDER BY sinif_adi"
    ).fetchall()]
    con.close()
    return rows


def _mevkiler_json() -> list[dict]:
    return [{"no": no, "emoji": emoji, "ad": ad} for no, emoji, ad in MEVKILER]


def _ogrenci_bul(ogrenci_id: int) -> dict | None:
    from database import _conn as _db
    con = _db()
    row = con.execute("""
        SELECT o.id, o.ad_soyad, o.ogr_no, o.sinif_id, s.sinif_adi,
               COUNT(t.id) AS tik_sayisi
        FROM ogrenciler o
        JOIN siniflar s ON s.id = o.sinif_id
        LEFT JOIN tik_kayitlari t ON t.ogrenci_id = o.id
        WHERE o.id = ?
        GROUP BY o.id
    """, (ogrenci_id,)).fetchone()
    con.close()
    if not row:
        return None
    return _ogrencilere_durum_ekle([dict(row)])[0]


def _ogrenci_no_ile_bul(ogr_no: int) -> dict | None:
    from database import _conn as _db
    con = _db()
    row = con.execute("""
        SELECT o.id, o.ad_soyad, o.ogr_no, o.sinif_id, s.sinif_adi,
               COUNT(t.id) AS tik_sayisi
        FROM ogrenciler o
        JOIN siniflar s ON s.id = o.sinif_id
        LEFT JOIN tik_kayitlari t ON t.ogrenci_id = o.id
        WHERE o.ogr_no = ?
        GROUP BY o.id
        ORDER BY s.sinif_adi, o.ad_soyad
        LIMIT 1
    """, (ogr_no,)).fetchone()
    con.close()
    if not row:
        return None
    return _ogrencilere_durum_ekle([dict(row)])[0]


def _ogrenci_rozetleri(ogrenci_id: int) -> list[dict]:
    from database import _conn as _db
    con = _db()
    rows = []
    for r in con.execute("""
        SELECT rozet_kodu, tarih
        FROM rozet_kayitlari
        WHERE ogrenci_id = ?
        ORDER BY tarih DESC
    """, (ogrenci_id,)).fetchall():
        d = dict(r)
        emoji, ad = rozet_emojileri_ve_metin(d["rozet_kodu"])
        d["emoji"] = emoji
        d["rozet_adi"] = ad
        rows.append(d)
    con.close()
    return rows


def _avatar(ogrenci: dict) -> dict:
    palette = ["#2563eb", "#16a34a", "#9333ea", "#dc2626", "#0891b2", "#ca8a04", "#be185d"]
    ad = ogrenci.get("ad_soyad", "?")
    initials = "".join(parca[:1] for parca in ad.split()[:2]).upper() or "?"
    renk = palette[int(ogrenci.get("id", 0)) % len(palette)]
    return {"initials": initials, "renk": renk}


def _kendi_ogrenci_id_veli_veya_ogrenci() -> int | None:
    if session.get("ogrenci_giris") and session.get("ogrenci_id"):
        return int(session["ogrenci_id"])
    if session.get("veli_ogrenci_id"):
        return int(session["veli_ogrenci_id"])
    return None


def _ogrenci_ozeti(ogrenci_id: int) -> dict | None:
    ogrenci = _ogrenci_bul(ogrenci_id)
    if not ogrenci:
        return None
    sinif_id = ogrenci["sinif_id"]
    sinif_liste = _ogrencilere_durum_ekle(sinif_ogrencileri(sinif_id))
    sirali = sorted(sinif_liste, key=lambda x: (x["tik_sayisi"], x["ad_soyad"]))
    disiplin_sira = next((i + 1 for i, o in enumerate(sirali) if o["id"] == ogrenci_id), None)
    gecmis = ogrenci_tik_gecmisi(ogrenci_id)
    odevler = ogrenci_odevleri(ogrenci_id)
    toplam_odev = len(odevler)
    tamamlanan_odev = sum(1 for o in odevler if o.get("tamamlandi"))
    kriter_sayim = {}
    for g in gecmis:
        kriter_sayim[g["kriter"]] = kriter_sayim.get(g["kriter"], 0) + 1
    en_sik_kriterler = sorted(
        [{"kriter": k, "sayi": v} for k, v in kriter_sayim.items()],
        key=lambda x: (-x["sayi"], x["kriter"])
    )[:6]
    son_gunler = {}
    for g in gecmis:
        gun = (g.get("tarih") or "")[:10]
        if gun:
            son_gunler[gun] = son_gunler.get(gun, 0) + 1
    gunluk_tikler = [{"gun": k, "tik": v} for k, v in sorted(son_gunler.items())[-10:]]
    stats = {
        "toplam_odev": toplam_odev,
        "tamamlanan_odev": tamamlanan_odev,
        "odev_orani": round(tamamlanan_odev * 100 / toplam_odev) if toplam_odev else 0,
        "rozet_sayisi": len(_ogrenci_rozetleri(ogrenci_id)),
        "olumlu_tik": ogrenci_olumlu_tik_sayisi(ogrenci_id),
        "gecmis_sayisi": len(gecmis),
        "temiz_mi": ogrenci["tik_sayisi"] == 0,
        "risk_yuzde": min(100, ogrenci["tik_sayisi"] * 8),
        "en_sik_kriterler": en_sik_kriterler,
        "gunluk_tikler": gunluk_tikler,
    }
    return {
        "ogrenci": ogrenci,
        "avatar": _avatar(ogrenci),
        "gecmis": gecmis,
        "rozetler": _ogrenci_rozetleri(ogrenci_id),
        "maclar": [],
        "sinif_tablo": None,
        "sezon": next((r for r in sezon_siralama() if r.get("sinif_id") == sinif_id), None),
        "disiplin_sira": disiplin_sira,
        "odevler": odevler,
        "istatistik": stats,
        "gelisim": gelisim_ozeti(ogrenci_id),
        "veli_haftalik": haftalik_veli_ozeti(ogrenci_id),
    }


def _bildirimleri_hazirla() -> list[dict]:
    bildirimler = []
    for talep in bekleyen_ogrenci_talepleri():
        bildirimler.append({
            "tur": "Ittifak",
            "renk": "cyan",
            "baslik": f"{talep['sinif1_adi']} + {talep['sinif2_adi']}",
            "detay": "Ogrencilerden gelen bekleyen ittifak talebi",
            "hedef": url_for("oduller"),
        })
    if session.get("ogretmen_id"):
        pending = ogretmen_onay_bekleyen_ogrenci_maclari(
            int(session["ogretmen_id"]),
            _toplu_sifirlamaya_izinli_mi(int(session["ogretmen_id"])),
        )
        if pending:
            bildirimler.append({
                "tur": "Ogrenci maci",
                "renk": "amber",
                "baslik": f"{len(pending)} mac onay bekliyor",
                "detay": "Ogrencilerin girdigi spor sonucu ogretmen onayi bekliyor",
                "hedef": url_for("ogretmen_ogrenci_maclari"),
            })
    for o in _ogrencilere_durum_ekle(tum_okul_ogrencileri()):
        if o["tik_sayisi"] >= OLUMSUZ_TIK_LIMIT:
            bildirimler.append({
                "tur": "Disiplin",
                "renk": "red",
                "baslik": f"{o['ad_soyad']} ({o['sinif_adi']})",
                "detay": f"{o['tik_sayisi']} tik ile {o['etiket'] or 'izlem'} seviyesinde",
                "hedef": url_for("dashboard"),
            })
    gorev = bugun_gorev()
    if gorev:
        bildirimler.append({
            "tur": "Gorev",
            "renk": "emerald",
            "baslik": gorev.get("gorev", "Gunluk gorev"),
            "detay": f"{gorev.get('puan', 0)} puanlik gunluk sinif gorevi",
            "hedef": url_for("oduller"),
        })
    return bildirimler[:50]


# ══════════════════════════════════════════════════════════════════════════
# Giris / Cikis
# ══════════════════════════════════════════════════════════════════════════

@app.route("/")
def index():
    return redirect(url_for("dashboard") if "ogretmen_id" in session else url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    ogretmenler = [o["ad_soyad"] for o in tum_ogretmenler()]
    hata, secili = None, ""

    if request.method == "POST":
        ad    = request.form.get("ad_soyad", "").strip()
        sifre = request.form.get("sifre", "").strip()
        secili = ad

        if not ad:
            hata = "Lutfen adinizi secin."
        elif not sifre:
            hata = "Sifre bos birakilamaz."
        elif not ogretmen_dogrula(ad, sifre):
            hata = "Hatali sifre! Sifreniz icin okul yonetimine basvurun."
        else:
            oid = ogretmen_id_bul(ad)
            session["ogretmen_id"] = oid
            session["ogretmen_adi"] = ad
            session["ogretmen_yetki"] = ogretmen_yetki_al(oid)
            if session.get("ogretmen_yetki") == "rapor":
                return redirect(url_for("rapor_ozet"))
            return redirect(url_for("dashboard"))

    return render_template("login.html", ogretmenler=ogretmenler, hata=hata, secili=secili)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/ogrenci/giris", methods=["GET", "POST"])
def ogrenci_giris():
    hata = None
    next_url = request.values.get("next") or url_for("ogrenci_gelisim_panel")
    if not next_url.startswith("/"):
        next_url = url_for("ogrenci_gelisim_panel")

    if session.get("ogrenci_giris"):
        return redirect(next_url)

    if request.method == "POST":
        sifre = request.form.get("sifre", "").strip()
        ogr_no = request.form.get("ogr_no", type=int)
        if sifre != OGRENCI_SIFRE:
            hata = "Ogrenci sifresi hatali."
        elif not ogr_no:
            hata = "Lutfen ogrenci numarasini girin."
        else:
            ogrenci = _ogrenci_no_ile_bul(ogr_no)
            if not ogrenci:
                hata = "Bu numarayla ogrenci bulunamadi."
            else:
                session["ogrenci_giris"] = True
                session["ogrenci_id"] = ogrenci["id"]
                return redirect(next_url)

    return render_template("ogrenci_login.html", hata=hata, next_url=next_url)


@app.route("/ogrenci/cikis")
def ogrenci_cikis():
    session.pop("ogrenci_giris", None)
    session.pop("ogrenci_id", None)
    return redirect(url_for("ogrenci_giris"))


@app.route("/ogrenci/ben")
@ogrenci_giris_zorunlu
def ogrenci_ben():
    return redirect(url_for("ogrenci_gelisim_panel"))


@app.route("/ogrenci/gelisim")
@ogrenci_giris_zorunlu
def ogrenci_gelisim_panel():
    oid = int(session["ogrenci_id"])
    o = _ogrenci_bul(oid)
    if not o:
        session.pop("ogrenci_giris", None)
        session.pop("ogrenci_id", None)
        return redirect(url_for("ogrenci_giris"))
    ozet = gelisim_ozeti(oid)
    gl = gelisim_ligi()
    sinif_satir = next((s for s in gl if s["sinif_id"] == o["sinif_id"]), None)
    sinif_sira = next((i for i, s in enumerate(gl, 1) if s["sinif_id"] == o["sinif_id"]), None)
    
    # Yeni eklendi: Özellikler
    puan_dict = ozet.get("puan") or {}
    toplam_xp = int(puan_dict.get("xp", 0))
    ozellikler = ogrenci_ozellikleri_getir(oid, toplam_xp)
    
    return render_template(
        "ogrenci_xp_panel.html",
        ogrenci=o,
        ozet=ozet,
        sinif_satir=sinif_satir,
        sinif_sira=sinif_sira,
        sinif_sayisi=len(gl),
        ozellikler=ozellikler
    )


@app.route("/oyunlar")
@ogrenci_giris_zorunlu
def oyunlar():
    oid = int(session["ogrenci_id"])
    o = _ogrenci_bul(oid)
    sinif_adi = (o or {}).get("sinif_adi", "")
    try:
        sinif_seviyesi = int(str(sinif_adi).split("/")[0].strip()[0])
    except Exception:
        sinif_seviyesi = 5
    return render_template(
        "oyunlar.html",
        sinif_seviyesi=sinif_seviyesi,
        sinif_adi=sinif_adi,
        quiz_dersler=QUIZ_DERSLER.get(sinif_seviyesi, []),
        ders_emoji=DERS_EMOJI,
    )


@app.route("/ogrenci/mac")
@ogrenci_giris_zorunlu
def ogrenci_mac_panel():
    oid = int(session["ogrenci_id"])
    o = _ogrenci_bul(oid)
    if not o:
        session.pop("ogrenci_giris", None)
        session.pop("ogrenci_id", None)
        return redirect(url_for("ogrenci_giris"))
    siniflar = [s for s in _tum_siniflar() if int(s["id"]) != int(o["sinif_id"])]
    maclar = ogrenci_mac_listesi(sinif_id=int(o["sinif_id"]), limit=30)
    return render_template(
        "ogrenci_mac.html",
        ogrenci=o,
        siniflar=siniflar,
        maclar=maclar,
    )


@app.route("/ogrenci/mac/olustur", methods=["POST"])
@ogrenci_giris_zorunlu
def ogrenci_mac_olustur_route():
    oid = int(session["ogrenci_id"])
    sonuc = ogrenci_mac_olustur(
        oid,
        request.form.get("rakip_sinif_id", type=int),
        request.form.get("skor1", type=int),
        request.form.get("skor2", type=int),
        request.form.get("aciklama", ""),
        request.form.get("spor", "futbol"),
    )
    if sonuc.get("ok"):
        flash("Mac sonucu ogretmen onayina gonderildi.", "success")
    else:
        flash(sonuc.get("sebep", "Mac kaydedilemedi."), "warning")
    return redirect(url_for("ogrenci_mac_panel"))


@app.route("/api/ogrenci/mac/simule", methods=["POST"])
@ogrenci_giris_zorunlu
def api_ogrenci_mac_simule():
    import random
    
    oid = int(session["ogrenci_id"])
    o = _ogrenci_bul(oid)
    if not o:
        return jsonify({"ok": False, "sebep": "Ogrenci bulunamadi"}), 404
        
    veri = request.get_json(silent=True) or {}
    rakip_sinif_id = veri.get("rakip_sinif_id")
    spor = (veri.get("spor", "futbol") or "futbol").strip().lower()
    if spor not in {"futbol", "voleybol"}:
        return jsonify({"ok": False, "sebep": "Spor turu gecersiz"})
    
    if not rakip_sinif_id:
        return jsonify({"ok": False, "sebep": "Rakip sinif secilmedi"})
    try:
        rakip_sinif_id = int(rakip_sinif_id)
    except Exception:
        return jsonify({"ok": False, "sebep": "Rakip sinif gecersiz"})
        
    bizim_sinif_id = int(o["sinif_id"])
    if rakip_sinif_id == bizim_sinif_id:
        return jsonify({"ok": False, "sebep": "Rakip sinif kendi sinifiniz olamaz"})
    
    bizim_taktik = spor_taktik_yukle(bizim_sinif_id, spor) if spor == "voleybol" else taktik_yukle(bizim_sinif_id)
    rakip_taktik = spor_taktik_yukle(rakip_sinif_id, spor) if spor == "voleybol" else taktik_yukle(rakip_sinif_id)
    
    bizim_k = sinif_ogrencileri(bizim_sinif_id)
    rakip_k = sinif_ogrencileri(rakip_sinif_id)
    
    def calc_power(taktik, kadro):
        oyuncu_verisi = (taktik or {}).get("oyuncular") or (taktik or {}).get("yerlesim")
        if not oyuncu_verisi:
            return 50
        total = 0
        count = 0
        for pid, _p in oyuncu_verisi.items():
            ogr = next((x for x in kadro if str(x["id"]) == str(pid)), None)
            if ogr:
                xp = ogr.get("tik_sayisi", 0)
                ovr = min(99, 50 + int(xp * 0.8))
                total += ovr
                count += 1
        return int(total / count) if count > 0 else 50
        
    bizim_guc = calc_power(bizim_taktik, bizim_k)
    rakip_guc = calc_power(rakip_taktik, rakip_k)
    
    bizim_skor = 0
    rakip_skor = 0
    
    anlatim = []
    anlatim.append(f"🏁 {o['sinif_adi']} vs Rakip ({spor.capitalize()} Simülasyonu)")
    anlatim.append(f"💪 Takım Gücü: Biz (⚡{bizim_guc}) - Rakip (⚡{rakip_guc})")
    anlatim.append("---")
    
    if spor == "voleybol":
        biz_set = 0
        rak_set = 0
        nokta_ihtimali = max(35, min(65, 50 + (bizim_guc - rakip_guc) // 2))
        for i in range(1, 6):
            if biz_set == 3 or rak_set == 3:
                break
            b_score = 0
            r_score = 0
            while (b_score < 25 and r_score < 25) or abs(b_score - r_score) < 2:
                if random.randint(1, 100) <= nokta_ihtimali:
                    b_score += 1
                else:
                    r_score += 1
                if b_score + r_score > 90:
                    if b_score == r_score:
                        if nokta_ihtimali >= 50:
                            b_score += 1
                        else:
                            r_score += 1
                    if abs(b_score - r_score) < 2:
                        if b_score > r_score:
                            b_score += 1
                        else:
                            r_score += 1
            if b_score > r_score:
                biz_set += 1
                anlatim.append(f"🏐 {i}. Set: {b_score}-{r_score} (Set bizim!)")
            else:
                rak_set += 1
                anlatim.append(f"🏐 {i}. Set: {b_score}-{r_score} (Set rakibin)")
        bizim_skor = biz_set
        rakip_skor = rak_set
    else:
        def poisson(lam):
            limit = 2.718281828459045 ** (-lam)
            carpim = 1.0
            adet = 0
            while carpim > limit:
                adet += 1
                carpim *= random.random()
            return max(0, adet - 1)

        fark = max(-2.0, min(2.0, (bizim_guc - rakip_guc) / 28))
        bizim_skor = min(7, poisson(1.25 + fark * 0.45))
        rakip_skor = min(7, poisson(1.25 - fark * 0.45))
        olaylar = ["biz"] * bizim_skor + ["rakip"] * rakip_skor
        random.shuffle(olaylar)
        dakikalar = sorted(random.sample(range(2, 90), len(olaylar))) if olaylar else []
        anlik_biz = 0
        anlik_rakip = 0
        for dk, takim in zip(dakikalar, olaylar):
            if takim == "biz":
                anlik_biz += 1
                anlatim.append(f"⚽ Dk {dk}: Hazirlanan atak gol oldu. ({anlik_biz}-{anlik_rakip})")
            else:
                anlik_rakip += 1
                anlatim.append(f"❌ Dk {dk}: Rakip boslugu buldu ve gol atti. ({anlik_biz}-{anlik_rakip})")
        bizim_skor = anlik_biz
        rakip_skor = anlik_rakip
        if not olaylar:
            anlatim.append("⚽ Iki takim da savunmada dikkatliydi, gol sesi cikmadi.")
                
    anlatim.append("---")
    if bizim_skor > rakip_skor:
        anlatim.append("🏆 MAÇ SONUCU: KAZANDIK!")
    elif bizim_skor < rakip_skor:
        anlatim.append("💀 MAÇ SONUCU: KAYBETTİK.")
    else:
        anlatim.append("🤝 MAÇ SONUCU: BERABERE.")
        
    return jsonify({
        "ok": True,
        "bizim_skor": bizim_skor,
        "rakip_skor": rakip_skor,
        "anlatim": "\\n".join(anlatim)
    })

@app.route("/api/ogrenci/mac/2d_data", methods=["POST"])
@ogrenci_giris_zorunlu
def api_ogrenci_mac_2d_data():
    oid = int(session["ogrenci_id"])
    o = _ogrenci_bul(oid)
    if not o:
        return jsonify({"ok": False}), 404
        
    veri = request.get_json(silent=True) or {}
    rakip_sinif_id = veri.get("rakip_sinif_id")
    spor = (veri.get("spor", "futbol") or "futbol").strip().lower()
    if spor not in {"futbol", "voleybol"}:
        return jsonify({"ok": False, "sebep": "Spor turu gecersiz"}), 400
    try:
        rakip_sinif_id = int(rakip_sinif_id)
    except Exception:
        return jsonify({"ok": False, "sebep": "Rakip sinif gecersiz"}), 400
    
    bizim_sinif_id = int(o["sinif_id"])
    if rakip_sinif_id == bizim_sinif_id:
        return jsonify({"ok": False, "sebep": "Rakip sinif kendi sinifiniz olamaz"}), 400
    
    bizim_taktik = spor_taktik_yukle(bizim_sinif_id, spor) if spor == "voleybol" else taktik_yukle(bizim_sinif_id)
    rakip_taktik = spor_taktik_yukle(rakip_sinif_id, spor) if spor == "voleybol" else taktik_yukle(rakip_sinif_id)
    
    biz_kadro = sinif_ogrencileri(bizim_sinif_id)
    rak_kadro = sinif_ogrencileri(rakip_sinif_id)
    
    def extract_players(taktik, kadro, takim_renk):
        players = []
        taktik_oyuncular = (taktik or {}).get("oyuncular") or (taktik or {}).get("yerlesim")
        if not taktik_oyuncular:
            if spor == "voleybol":
                default_positions = [
                    {"x": 25, "y": 58}, {"x": 50, "y": 58}, {"x": 75, "y": 58},
                    {"x": 25, "y": 82}, {"x": 50, "y": 82}, {"x": 75, "y": 82}
                ]
                limit = 6
            else:
                default_positions = [
                    {"x": 50, "y": 90}, # GK
                    {"x": 20, "y": 70}, {"x": 40, "y": 75}, {"x": 60, "y": 75}, {"x": 80, "y": 70}, # DEF
                    {"x": 20, "y": 50}, {"x": 40, "y": 50}, {"x": 60, "y": 50}, {"x": 80, "y": 50}, # MID
                    {"x": 35, "y": 30}, {"x": 65, "y": 30} # FWD
                ]
                limit = 11
            for i, ogr in enumerate(kadro[:limit]):
                pos = default_positions[i] if i < len(default_positions) else {"x": 50, "y": 50}
                ozet = gelisim_ozeti(ogr["id"])
                puan_dict = ozet.get("puan") or {}
                xp = int(puan_dict.get("xp", 0))
                ovr = min(99, 50 + int(xp * 0.8))
                
                oz = ogrenci_ozellikleri_getir(ogr["id"], xp)
                
                players.append({
                    "id": ogr["id"],
                    "ad": ogr["ad_soyad"].split()[0],
                    "numara": i + 1,
                    "rol": "Oyuncu",
                    "baseX": pos["x"],
                    "baseY": pos["y"],
                    "ovr": ovr,
                    "renk": takim_renk,
                    "ozellikler": oz
                })
            return players
            
        for pid, p in taktik_oyuncular.items():
            ogr = next((x for x in kadro if str(x["id"]) == str(pid)), None)
            if ogr:
                ozet = gelisim_ozeti(ogr["id"])
                puan_dict = ozet.get("puan") or {}
                xp = int(puan_dict.get("xp", 0))
                ovr = min(99, 50 + int(xp * 0.8))
                
                oz = ogrenci_ozellikleri_getir(ogr["id"], xp)
                
                players.append({
                    "id": ogr["id"],
                    "ad": ogr["ad_soyad"].split()[0],
                    "numara": p.get("mevki_no", 10),
                    "rol": p.get("rol", ""),
                    "baseX": p.get("x", 50),
                    "baseY": p.get("y", 50),
                    "ovr": ovr,
                    "renk": taktik.get("renk", takim_renk),
                    "ozellikler": oz
                })
        return players
        
    bizim_oyuncular = extract_players(bizim_taktik, biz_kadro, "#3b82f6")
    rakip_oyuncular = extract_players(rakip_taktik, rak_kadro, "#ef4444")
    
    return jsonify({
        "ok": True,
        "bizim_takim": {"ad": o["sinif_adi"], "oyuncular": bizim_oyuncular},
        "rakip_takim": {"ad": "Rakip", "oyuncular": rakip_oyuncular}
    })

@app.route("/api/ogrenci/ozellik/artir", methods=["POST"])
@ogrenci_giris_zorunlu
def api_ogrenci_ozellik_artir():
    oid = int(session["ogrenci_id"])
    ozet = gelisim_ozeti(oid)
    puan_dict = ozet.get("puan") or {}
    toplam_xp = int(puan_dict.get("xp", 0))
    veri = request.get_json(silent=True) or {}
    ozellik = veri.get("ozellik")
    
    sonuc = ogrenci_ozellik_artir(oid, ozellik, toplam_xp)
    return jsonify(sonuc)

@app.route("/ogrenci/mac/2d")
@ogrenci_giris_zorunlu
def ogrenci_mac_saha_2d():
    return render_template("mac_2d.html")

@app.route("/ogrenci/voleybol/2d")
@ogrenci_giris_zorunlu
def ogrenci_voleybol_saha_2d():
    return render_template("voleybol_2d.html")


@app.route("/ogrenci/taktik")
@ogrenci_giris_zorunlu
def ogrenci_taktik():
    oid = int(session["ogrenci_id"])
    o = _ogrenci_bul(oid)
    if not o:
        return redirect(url_for("ogrenci_giris"))
    sinif_id = int(o["sinif_id"])
    return render_template(
        "taktik.html",
        sinif_id=sinif_id,
        sinif_adi=o["sinif_adi"],
        kadro=sinif_ogrencileri(sinif_id),
        kayit=taktik_yukle(sinif_id),
        mevkiler_json=_mevkiler_json(),
        ana_menu_url=url_for("ogrenci_mac_panel"),
        kaydet_url=url_for("api_ogrenci_taktik_kaydet"),
        ben_id=oid,
    )


@app.route("/api/ogrenci/taktik/kaydet", methods=["POST"])
@ogrenci_giris_zorunlu
def api_ogrenci_taktik_kaydet():
    oid = int(session["ogrenci_id"])
    o = _ogrenci_bul(oid)
    if not o:
        return jsonify({"ok": False, "sebep": "Ogrenci bulunamadi"}), 404
    veri = request.get_json(silent=True) or {}
    return jsonify(taktik_kaydet(int(o["sinif_id"]), json.dumps(veri, ensure_ascii=False)))


@app.route("/ogrenci/voleybol-taktik")
@ogrenci_giris_zorunlu
def ogrenci_voleybol_taktik():
    oid = int(session["ogrenci_id"])
    o = _ogrenci_bul(oid)
    if not o:
        return redirect(url_for("ogrenci_giris"))
    sinif_id = int(o["sinif_id"])
    return render_template(
        "voleybol_taktik.html",
        sinif_id=sinif_id,
        sinif_adi=o["sinif_adi"],
        kadro=sinif_ogrencileri(sinif_id),
        kayit=spor_taktik_yukle(sinif_id, "voleybol"),
        ana_menu_url=url_for("ogrenci_mac_panel"),
        kaydet_url=url_for("api_ogrenci_voleybol_taktik_kaydet"),
    )


@app.route("/api/ogrenci/voleybol-taktik/kaydet", methods=["POST"])
@ogrenci_giris_zorunlu
def api_ogrenci_voleybol_taktik_kaydet():
    oid = int(session["ogrenci_id"])
    o = _ogrenci_bul(oid)
    if not o:
        return jsonify({"ok": False, "sebep": "Ogrenci bulunamadi"}), 404
    veri = request.get_json(silent=True) or {}
    return jsonify(spor_taktik_kaydet(int(o["sinif_id"]), "voleybol", json.dumps(veri, ensure_ascii=False)))


@app.route("/api/oyun/quiz-sorular")
@ogrenci_giris_zorunlu
def oyun_quiz_sorular_api():
    import random

    oid = int(session["ogrenci_id"])
    o = _ogrenci_bul(oid)
    sinif_adi = (o or {}).get("sinif_adi", "")
    try:
        sinif_seviyesi = int(str(sinif_adi).split("/")[0].strip()[0])
    except Exception:
        sinif_seviyesi = request.args.get("sinif_seviyesi", 5, type=int)
    ders = (request.args.get("ders") or "").strip()
    adet = max(8, min(request.args.get("adet", 40, type=int), 80))
    quiz_sorulari_yukle()
    from database import _conn as _db_conn
    from database import _quiz_init

    con = _db_conn()
    _quiz_init(con)
    params: list = [sinif_seviyesi]
    where = "sinif_seviyesi=?"
    if ders:
        where += " AND ders=?"
        params.append(ders)
    rows = [dict(r) for r in con.execute(
        f"SELECT * FROM quiz_sorulari WHERE {where} ORDER BY RANDOM() LIMIT ?",
        (*params, adet),
    ).fetchall()]
    con.close()
    sorular = []
    for r in rows:
        dogru = str(r.get("dogru_cevap") or "").upper()
        secenekler = [
            {"harf": "A", "metin": r.get("secenek_a") or ""},
            {"harf": "B", "metin": r.get("secenek_b") or ""},
            {"harf": "C", "metin": r.get("secenek_c") or ""},
            {"harf": "D", "metin": r.get("secenek_d") or ""},
        ]
        random.shuffle(secenekler)
        dogru_metin = next((s["metin"] for s in secenekler if s["harf"] == dogru), "")
        sorular.append({
            "id": r["id"],
            "ders": r.get("ders") or "",
            "soru": r.get("soru") or "",
            "secenekler": secenekler,
            "dogru": dogru,
            "dogru_metin": dogru_metin,
        })
    return jsonify({"ok": True, "sinif_seviyesi": sinif_seviyesi, "sorular": sorular})


@app.route("/api/oyun/puan", methods=["POST"])
@ogrenci_giris_zorunlu
def oyun_puan_api():
    veri = request.get_json(silent=True) or {}
    oid = int(session["ogrenci_id"])
    sonuc = oyun_puani_kaydet(
        oid,
        veri.get("oyun") or "Oyun",
        int(veri.get("puan") or 0),
    )
    o = _ogrenci_bul(oid)
    ek = {}
    if o:
        ek["sinif_adi"] = o.get("sinif_adi", "")
    return jsonify({**sonuc, **ek})


@app.route("/api/ogrenci/gorev-tamamla", methods=["POST"])
@ogrenci_giris_zorunlu
def api_ogrenci_gorev_tamamla():
    return jsonify(gelisim_gorev_tamamla(int(session["ogrenci_id"])))


@app.route("/gelisim")
@ogrenci_giris_zorunlu
def gelisim_merkezi():
    return redirect(url_for("ogrenci_gelisim_panel"))


@app.route("/patika")
@ogrenci_giris_zorunlu
def ogrenci_patika():
    return redirect(url_for("ogrenci_gelisim_panel"))


@app.route("/gelisim/gorev-tamamla", methods=["POST"])
@ogrenci_giris_zorunlu
def gelisim_gorev_tamamla_route():
    sonuc = gelisim_gorev_tamamla(int(session["ogrenci_id"]))
    if sonuc.get("ok"):
        flash("Görev tamamlandı; XP ve sınıf lig katkın kaydedildi.", "success")
    else:
        flash(sonuc.get("sebep", "İşlem yapılamadı"), "error")
    return redirect(url_for("ogrenci_gelisim_panel"))


@app.route("/gelisim/sandik-ac", methods=["POST"])
@ogrenci_giris_zorunlu
def gelisim_sandik_ac_route():
    return redirect(url_for("ogrenci_gorunum"))


@app.route("/gelisim/telafi", methods=["POST"])
@ogrenci_giris_zorunlu
def gelisim_telafi_route():
    return redirect(url_for("ogrenci_gorunum"))


@app.route("/gelisim/tebrik", methods=["POST"])
@ogrenci_giris_zorunlu
def gelisim_tebrik_route():
    return redirect(url_for("ogrenci_gorunum"))


@app.route("/pazar")
@ogrenci_giris_zorunlu
def pazar():
    oid = int(session["ogrenci_id"])
    o = _ogrenci_bul(oid)
    if not o:
        session.pop("ogrenci_giris", None)
        session.pop("ogrenci_id", None)
        return redirect(url_for("ogrenci_giris"))
    return render_template(
        "pazar.html",
        ogrenci=o,
        urunler=pazar_urunleri_ogrenci(oid),
        gelisim=gelisim_ozeti(oid),
    )


@app.route("/pazar/satin-al", methods=["POST"])
@ogrenci_giris_zorunlu
def pazar_satin_al_route():
    kod = request.form.get("urun_kodu", "").strip()
    sonuc = pazar_satin_al(int(session["ogrenci_id"]), kod)
    if sonuc.get("ok"):
        flash(f"Satın alındı: {sonuc['urun']['ad']}", "success")
    else:
        flash(sonuc.get("sebep", "İşlem yapılamadı"), "error")
    return redirect(url_for("pazar"))


@app.route("/envanter")
@ogrenci_giris_zorunlu
def envanter():
    oid = int(session["ogrenci_id"])
    o = _ogrenci_bul(oid)
    if not o:
        session.pop("ogrenci_giris", None)
        session.pop("ogrenci_id", None)
        return redirect(url_for("ogrenci_giris"))
    return render_template(
        "envanter.html",
        ogrenci=o,
        envanter=envanter_listele(oid),
    )


@app.route("/envanter/aktif", methods=["POST"])
@ogrenci_giris_zorunlu
def envanter_aktif_route():
    oid = int(session["ogrenci_id"])
    kod = request.form.get("urun_kodu", "").strip()
    sonuc = envanter_aktif_ayarla(oid, kod)
    if sonuc.get("ok"):
        flash("Gorunen envanter guncellendi.", "success")
    else:
        flash(sonuc.get("sebep", "Islem yapilamadi"), "error")
    return redirect(url_for("envanter"))


@app.route("/karne")
@ogrenci_giris_zorunlu
def ogrenci_karne():
    return redirect(url_for("ogrenci_gorunum"))


@app.route("/veli/karne")
def veli_karne():
    return redirect(url_for("veli_panel"))


@app.route("/ogretmen/merkez")
@giris_zorunlu
def ogretmen_merkez():
    return render_template("ogretmen_merkez.html",
                           merkez=ogretmen_bildirim_merkezi(),
                           sablonlar=GOREV_SABLONLARI)


@app.route("/ogretmen/not", methods=["POST"])
@giris_zorunlu
def ogretmen_not_route():
    ogrenci_id = request.form.get("ogrenci_id", type=int)
    not_metni = request.form.get("not_metni", "")
    if ogrenci_id:
        ogretmen_notu_ekle(ogrenci_id, session["ogretmen_id"], not_metni,
                           request.form.get("veliye_acik") == "1")
    return redirect(url_for("ogretmen_merkez"))


@app.route("/gelisim-ligi")
@giris_zorunlu
def gelisim_ligi_sayfa():
    return redirect(url_for("dashboard"))


@app.route("/veli/giris", methods=["GET", "POST"])
def veli_giris():
    hata = None
    if request.method == "POST":
        sifre = request.form.get("sifre", "").strip()
        ogr_no = request.form.get("ogr_no", type=int)
        if sifre != VELI_SIFRE:
            hata = "Veli sifresi hatali."
        elif not ogr_no:
            hata = "Lutfen ogrenci numarasini girin."
        else:
            ogrenci = _ogrenci_no_ile_bul(ogr_no)
            if not ogrenci:
                hata = "Bu numarayla ogrenci bulunamadi."
            else:
                session["veli_ogrenci_id"] = ogrenci["id"]
                return redirect(url_for("veli_panel"))
    return render_template("veli_login.html", hata=hata)


@app.route("/veli")
def veli_panel():
    ogrenci_id = session.get("veli_ogrenci_id")
    if not ogrenci_id:
        return redirect(url_for("veli_giris"))
    o = _ogrenci_bul(int(ogrenci_id))
    if not o:
        session.pop("veli_ogrenci_id", None)
        return redirect(url_for("veli_giris"))
    gecmis = ogrenci_tik_gecmisi(int(ogrenci_id))
    odevler = ogrenci_odevleri(int(ogrenci_id), 40)
    return render_template(
        "tik_gecmisi.html",
        ogrenci=o,
        gecmis=gecmis,
        avatar=_avatar(o),
        veli_modu=True,
        odevler=odevler,
    )


@app.route("/veli/cikis")
def veli_cikis():
    session.pop("veli_ogrenci_id", None)
    return redirect(url_for("veli_giris"))


@app.route("/admin/sifreler", methods=["GET", "POST"])
def admin_sifreler():
    hata, liste = None, None
    if request.method == "POST":
        if request.form.get("admin_sifre") == ADMIN_SIFRE:
            liste = tum_sifre_listesi()
        else:
            hata = "Yonetici sifresi hatali."
    son_yedek = admin_meta_get("son_yedek_hatirlatma", "")
    return render_template(
        "admin_sifreler.html",
        liste=liste,
        hata=hata,
        son_yedek=son_yedek,
    )


# ══════════════════════════════════════════════════════════════════════════
# Ana Panel
# ══════════════════════════════════════════════════════════════════════════

@app.route("/bilgilendirme/sil/<int:bilgi_id>", methods=["POST"])
@giris_zorunlu
def bilgilendirme_sil(bilgi_id: int):
    if not bilgilendirme_yetkili_mi():
        abort(403)
    ad = session.get("ogretmen_adi", "") or ""
    sonuc = bilgilendirme_yayinlayan_icin_sil(bilgi_id, ad)
    if sonuc.get("sebep") == "yetkisiz":
        abort(403)
    if sonuc["ok"]:
        flash("Duyuru kaldırıldı.", "success")
    elif sonuc.get("sebep") == "bulunamadi":
        flash("Duyuru bulunamadı.", "error")
    elif sonuc.get("sebep") == "zaten_kaldirilmis":
        flash("Bu duyuru zaten kaldırılmış.", "info")
    return redirect(url_for("bilgilendirme_yonetimi"))


@app.route("/bilgilendirme", methods=["GET", "POST"])
@giris_zorunlu
def bilgilendirme_yonetimi():
    hedefler = [
        ("herkes", "Herkes"),
        ("ogretmen", "Öğretmenler"),
        ("ogrenci", "Öğrenciler"),
        ("veli", "Veliler"),
    ]
    hedef_adlari = dict(hedefler)
    if not bilgilendirme_yetkili_mi():
        return render_template(
            "bilgilendirme.html",
            yetki_yok=True,
            hedefler=hedefler,
            hedef_adlari=hedef_adlari,
            duyurular=bilgilendirme_listesi(10),
            ogretmen_adi="",
        ), 403

    hata = None
    basari = None
    secili_hedef = request.form.get("hedef", "herkes")
    baslik = request.form.get("baslik", "").strip()
    metin = request.form.get("metin", "").strip()

    if request.method == "POST":
        if not baslik:
            hata = "Başlık boş bırakılamaz."
        elif not metin:
            hata = "Bilgilendirme metni boş bırakılamaz."
        else:
            bilgilendirme_ekle(
                baslik=baslik,
                metin=metin,
                hedef=secili_hedef,
                yayinlayan=session.get("ogretmen_adi", "Yönetim"),
            )
            basari = "Bilgilendirme yayınlandı. Seçilen kişiler bir sonraki ana ekran açılışında görecek."
            baslik = ""
            metin = ""

    return render_template(
        "bilgilendirme.html",
        yetki_yok=False,
        hedefler=hedefler,
        hedef_adlari=hedef_adlari,
        secili_hedef=secili_hedef,
        baslik=baslik,
        metin=metin,
        hata=hata,
        basari=basari,
        duyurular=bilgilendirme_listesi(20),
        ogretmen_adi=session.get("ogretmen_adi", "") or "",
    )


@app.route("/dashboard")
@giris_zorunlu
def dashboard():
    ogretmen_id  = session["ogretmen_id"]
    ogretmen_adi = session.get("ogretmen_adi", "")
    siniflar     = ogretmen_siniflari(ogretmen_id)

    if not siniflar:
        return render_template("dashboard.html", siniflar=[], aktif=None,
                               ogrenciler=[], kriterler=KRITERLER,
                               olumlu_kriterler=OLUMLU_KRITERLER,
                               olumlu_satirlari=[],
                               yetkili_popup=False,
                               bugunki_mufettis=None,
                               bekleyen_talepler=[],
                               tum_sinif_ogrencileri_popup=[],
                               toplu_sifirlamaya_izin=_toplu_sifirlamaya_izinli_mi(ogretmen_id))

    try:
        aktif_id = int(request.args.get("sinif", siniflar[0]["id"]))
    except (TypeError, ValueError):
        aktif_id = siniflar[0]["id"]

    aktif      = next((s for s in siniflar if s["id"] == aktif_id), siniflar[0])
    ogrenciler = _ogrencilere_durum_ekle(sinif_ogrencileri(aktif["id"]))

    yetkili = True
    bugunki_muf = bugun_mufettis()
    bek_talepler = bekleyen_ogrenci_talepleri()

    popup_ogrenciler = []
    if yetkili and not bugunki_muf:
        from database import _conn as _db
        _con = _db()
        _siniflar = [dict(r) for r in _con.execute(
            "SELECT id, sinif_adi FROM siniflar ORDER BY sinif_adi"
        ).fetchall()]
        _con.close()
        for s in _siniflar:
            olist = sinif_ogrencileri(s["id"])
            for o in olist:
                popup_ogrenciler.append({
                    "id": o["id"],
                    "ad_soyad": o["ad_soyad"],
                    "sinif_adi": s["sinif_adi"],
                    "sinif_id": s["id"],
                })

    olumlu_satirlari = sinif_olumlu_gecmis(aktif["id"])
    ids_ogr = [o["id"] for o in ogrenciler]
    olumlu_h = ogrenci_olumlu_tik_sayilari(ids_ogr)
    roz_harita = ogrenci_rozetleri_yayin_map(ids_ogr, limit=6)
    env_harita = ogrenci_aktif_envanter_map(ids_ogr)
    for o in ogrenciler:
        o["rozetler"] = roz_harita.get(o["id"], [])
        o["aktif_envanter"] = env_harita.get(o["id"])
        o["olumlu_tik"] = olumlu_h.get(o["id"], 0)

    return render_template("dashboard.html",
                           siniflar=siniflar, aktif=aktif,
                           ogrenciler=ogrenciler,
                           kriterler=KRITERLER,
                           olumlu_kriterler=OLUMLU_KRITERLER,
                           olumlu_satirlari=olumlu_satirlari,
                           yetkili_popup=yetkili,
                           bugunki_mufettis=bugunki_muf,
                           bekleyen_talepler=bek_talepler,
                           tum_sinif_ogrencileri_popup=popup_ogrenciler,
                           toplu_sifirlamaya_izin=_toplu_sifirlamaya_izinli_mi(ogretmen_id))


@app.route("/api/sinif/<int:sinif_id>")
@giris_zorunlu
def api_sinif(sinif_id):
    if not _ogretmen_sinifinda_mi(session["ogretmen_id"], sinif_id):
        return jsonify({"ok": False, "sebep": "Yetkisiz"}), 403
    liste = _ogrencilere_durum_ekle(sinif_ogrencileri(sinif_id))
    ids = [o["id"] for o in liste]
    olumlu_h = ogrenci_olumlu_tik_sayilari(ids)
    roz_harita = ogrenci_rozetleri_yayin_map(ids, limit=6)
    env_harita = ogrenci_aktif_envanter_map(ids)
    for o in liste:
        o["rozetler"] = roz_harita.get(o["id"], [])
        o["aktif_envanter"] = env_harita.get(o["id"])
        o["olumlu_tik"] = olumlu_h.get(o["id"], 0)
    return jsonify(liste)


@app.route("/bildirimler")
@giris_zorunlu
def bildirimler():
    return render_template("bildirimler.html", bildirimler=_bildirimleri_hazirla())


@app.route("/rapor/ozet")
@giris_zorunlu
def rapor_ozet():
    siniflar = _tum_siniflar()
    sinif_idleri = [s["id"] for s in siniflar]
    okul = _ogrencilere_durum_ekle(tum_okul_ogrencileri())
    sinif_ozetleri = []
    toplam_tik = sum(o["tik_sayisi"] for o in okul)
    ortalama_tik = round(toplam_tik / len(okul), 2) if okul else 0
    heatmap_okul = []
    heatmap_siniflar = []
    for s in siniflar:
        liste = _ogrencilere_durum_ekle(sinif_ogrencileri(s["id"]))
        sinif_tik = sum(o["tik_sayisi"] for o in liste)
        ogrenci_sayisi = len(liste)
        ortalama = round(sinif_tik / ogrenci_sayisi, 2) if ogrenci_sayisi else 0
        dagilim = {
            "temiz": sum(1 for o in liste if o["tik_sayisi"] == 0),
            "uyari": sum(1 for o in liste if 1 <= o["tik_sayisi"] <= 2),
            "veli": sum(1 for o in liste if 6 <= o["tik_sayisi"] <= 8),
            "tutanak": sum(1 for o in liste if 9 <= o["tik_sayisi"] <= 11),
            "disiplin": sum(1 for o in liste if o["tik_sayisi"] >= 12),
        }
        sinif_ozetleri.append({
            "id": s["id"],
            "sinif_adi": s["sinif_adi"],
            "ogrenci": ogrenci_sayisi,
            "tik": sinif_tik,
            "idari": sum(1 for o in liste if o["tik_sayisi"] >= OLUMSUZ_TIK_LIMIT),
            "temiz": sum(1 for o in liste if o["tik_sayisi"] == 0),
            "ortalama": ortalama,
            "dagilim": dagilim,
        })
        for o in liste:
            heatmap_okul.append({
                "ad": o["ad_soyad"],
                "sinif": s["sinif_adi"],
                "tik": o["tik_sayisi"],
                "durum": o["durum"],
                "etiket": o["etiket"] or "Temiz",
            })
        hucreler = []
        for o in sorted(liste, key=lambda x: (-x["tik_sayisi"], x["ad_soyad"]))[:10]:
            sicaklik = min(100, o["tik_sayisi"] * 9)
            hucreler.append({
                "ad": o["ad_soyad"],
                "tik": o["tik_sayisi"],
                "etiket": o["etiket"] or "Temiz",
                "sicaklik": sicaklik,
            })
        if hucreler:
            heatmap_siniflar.append({"sinif_adi": s["sinif_adi"], "hucreler": hucreler})
    heatmap_okul.sort(key=lambda x: (-x["tik"], x["ad"]))
    heatmap_okul = heatmap_okul[:42]

    tik_satirlari = tik_kayitlari_siniflarda(sinif_idleri)
    kriter_pairs = kriter_dagilimi_satirlardan(tik_satirlari)[:14]
    kriter_chart = {
        "labels": [k[0][:48] for k in kriter_pairs],
        "values": [k[1] for k in kriter_pairs],
    }
    aylik_raw = aylik_tik_sayilari(tik_satirlari)
    aylik_chart = {
        "labels": [x["etiket"] for x in aylik_raw],
        "values": [x["adet"] for x in aylik_raw],
    }

    durum_dagilimi = durum_dagilimi_hesapla(okul)
    durum_chart = {
        "labels": ["Temiz", "Uyarı", "İdari", "Veli", "Tutanak", "Disiplin"],
        "values": [
            durum_dagilimi["temiz"],
            durum_dagilimi["uyari"],
            durum_dagilimi["idari"],
            durum_dagilimi["veli"],
            durum_dagilimi["tutanak"],
            durum_dagilimi["disiplin"],
        ],
    }
    sey = sinif_en_yuksek_ortalama_bul(sinif_ozetleri)
    oneriler = oneriler_derle(
        len(okul),
        toplam_tik,
        ortalama_tik,
        durum_dagilimi,
        kriter_pairs[:10],
        sey,
    )

    # Sınıf × yoğunluk: ortalama tik ile basit ısı ölçeği
    max_o = max((s["ortalama"] for s in sinif_ozetleri), default=0) or 1
    sinif_isi_skala = []
    for s in sinif_ozetleri:
        oran = min(100, int(100 * float(s["ortalama"]) / max_o))
        sinif_isi_skala.append({
            "sinif_adi": s["sinif_adi"],
            "ortalama": s["ortalama"],
            "tik": s["tik"],
            "isi": oran,
        })

    return render_template(
        "rapor_ozet.html",
        sinif_ozetleri=sinif_ozetleri,
        okul=okul,
        toplam_tik=toplam_tik,
        ortalama_tik=ortalama_tik,
        durum_dagilimi=durum_dagilimi,
        durum_chart=durum_chart,
        kriter_chart=kriter_chart,
        aylik_chart=aylik_chart,
        heatmap_okul=heatmap_okul,
        heatmap_siniflar=heatmap_siniflar,
        sinif_isi_skala=sinif_isi_skala,
        oneriler=oneriler,
        sinif_xp_siralama=gelisim_ligi(),
        sezon=sezon_siralama(),
        rozetler=son_rozetler(12),
        benim_siniflar=ogretmen_siniflari(session["ogretmen_id"]),
    )


@app.route("/rapor/ozet.csv")
@giris_zorunlu
def rapor_ozet_csv():
    satirlar = ["sinif,ogrenci,tik,idari,temiz"]
    for s in _tum_siniflar():
        liste = _ogrencilere_durum_ekle(sinif_ogrencileri(s["id"]))
        satirlar.append(
            f"{s['sinif_adi']},{len(liste)},{sum(o['tik_sayisi'] for o in liste)},"
            f"{sum(1 for o in liste if o['tik_sayisi'] >= OLUMSUZ_TIK_LIMIT)},"
            f"{sum(1 for o in liste if o['tik_sayisi'] == 0)}"
        )
    resp = make_response("\n".join(satirlar))
    resp.headers["Content-Type"] = "text/csv; charset=utf-8"
    resp.headers["Content-Disposition"] = "attachment; filename=okul-ozet-raporu.csv"
    return resp


@app.route("/turnuva")
@giris_zorunlu
def turnuva():
    return redirect(url_for("dashboard"))


@app.route("/yoklama")
@giris_zorunlu
def yoklama():
    siniflar = ogretmen_siniflari(session["ogretmen_id"])
    aktif_id = request.args.get("sinif", type=int) or (siniflar[0]["id"] if siniflar else 0)
    aktif = next((s for s in siniflar if s["id"] == aktif_id), siniflar[0] if siniflar else None)
    ogrenciler = sinif_ogrencileri(aktif["id"]) if aktif else []
    odevler = sinif_odevleri(aktif["id"]) if aktif else []
    secili_odev_id = request.args.get("odev", type=int) or (odevler[0]["id"] if odevler else 0)
    secili_odev = odev_detay_db(secili_odev_id) if secili_odev_id else None
    if secili_odev and aktif and secili_odev["sinif_id"] != aktif["id"]:
        secili_odev = None
    return render_template("yoklama.html", siniflar=siniflar, aktif=aktif,
                           ogrenciler=ogrenciler, odevler=odevler,
                           secili_odev=secili_odev)


@app.route("/odev/ekle", methods=["POST"])
@giris_zorunlu
def odev_olustur():
    sinif_id = request.form.get("sinif_id", type=int)
    sonuc = odev_ekle(
        sinif_id=sinif_id,
        ogretmen_id=session["ogretmen_id"],
        baslik=request.form.get("baslik", ""),
        aciklama=request.form.get("aciklama", ""),
        son_tarih=request.form.get("son_tarih", ""),
        ders=request.form.get("ders", "Genel"),
    )
    if not sonuc.get("ok"):
        return redirect(url_for("yoklama", sinif=sinif_id))
    flash(
        "Ödev yayınlandı. Öğrenci ve veli giriş yaptığında bildirim penceresinde görecek.",
        "success",
    )
    return redirect(url_for("yoklama", sinif=sinif_id, odev=sonuc["odev_id"]))


@app.route("/odev/<int:odev_id>/tamamla/<int:ogrenci_id>", methods=["POST"])
@giris_zorunlu
def odev_tamamla_route(odev_id, ogrenci_id):
    isaretli = request.form.get("tamamlandi") == "1"
    if isaretli:
        odev_tamamla(odev_id, ogrenci_id, session["ogretmen_id"])
    else:
        odev_tamamlandi_kaldir(odev_id, ogrenci_id)
    detay = odev_detay_db(odev_id)
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return jsonify({"ok": True})
    sinif_id = detay["sinif_id"] if detay else request.form.get("sinif_id", "")
    return redirect(url_for("yoklama", sinif=sinif_id, odev=odev_id))


@app.route("/odevler", methods=["GET", "POST"])
@giris_zorunlu
def odevler():
    oid = session["ogretmen_id"]
    from database import odev_olustur, odevleri_getir

    if request.method == "POST":
        sinif_id = request.form.get("sinif_id", type=int)
        ders_adi = request.form.get("ders_adi", "").strip()
        tema_adi = request.form.get("tema_adi", "").strip()
        konu_adi = request.form.get("konu_adi", "").strip()
        if not konu_adi:
            konu_adi = request.form.get("konu_adi_ek", "").strip()
        else:
            ek = request.form.get("konu_adi_ek", "").strip()
            if ek and ek != konu_adi:
                konu_adi = konu_adi + " — " + ek
        og_json = (request.form.get("ogrenme_ciktilari_json") or "").strip() or "[]"
        try:
            json.loads(og_json)
        except Exception:
            og_json = "[]"

        sev = request.form.get("sinif_seviyesi", type=int)
        if sev not in (5, 6, 7, 8):
            sev = None

        if sinif_id and ders_adi and tema_adi:
            odev_id = odev_olustur(
                oid,
                sinif_id,
                ders_adi,
                tema_adi,
                konu_adi,
                og_json,
                sinif_seviyesi=sev,
            )
            return redirect(url_for("odev_tema_detay", odev_id=odev_id))

    odev_listesi = odevleri_getir(oid)
    return render_template(
        "odevler.html",
        odevler=odev_listesi,
        siniflar=ogretmen_siniflari(oid),
        yetki=session.get("ogretmen_yetki"),
    )


@app.route("/odev/<int:odev_id>", methods=["GET", "POST"])
@giris_zorunlu
def odev_tema_detay(odev_id):
    oid = session["ogretmen_id"]
    from database import odev_detay_getir, odev_durum_guncelle

    detay = odev_detay_getir(odev_id, oid)
    if not detay:
        return redirect(url_for("odevler"))

    try:
        detay["odev"]["ogrenme_ciktilari"] = json.loads(
            detay["odev"].get("ogrenme_ciktilari_json") or "[]"
        )
    except Exception:
        detay["odev"]["ogrenme_ciktilari"] = []
    try:
        detay["odev"]["ogrenme_cikti_kodlari"] = json.loads(
            detay["odev"].get("ogrenme_cikti_kodlari_json") or "[]"
        )
    except Exception:
        detay["odev"]["ogrenme_cikti_kodlari"] = []

    if request.method == "POST":
        for key, val in request.form.items():
            if key.startswith("durum_"):
                ogr_id = int(key.split("_")[1])
                odev_durum_guncelle(odev_id, ogr_id, val)
        return redirect(url_for("odev_tema_detay", odev_id=odev_id))

    return render_template(
        "odev_detay.html",
        odev=detay["odev"],
        ogrenciler=detay["ogrenciler"],
        yetki=session.get("ogretmen_yetki"),
    )


@app.route("/api/curriculum/temel-egitim")
@giris_zorunlu
def api_curriculum_temel_egitim():
    """TYMM temel egitim ozeti: tema > konu > ogrenme ciktisi (data/meb_temel_egitim_curriculum.json)."""
    path = os.path.join(app.root_path, "data", "meb_temel_egitim_curriculum.json")
    try:
        with open(path, encoding="utf-8") as f:
            return jsonify(json.load(f))
    except OSError:
        return jsonify({"hata": "Mufredat dosyasi bulunamadi", "dersler": {}}), 404


@app.route("/api/odev/mufredat-ozet")
@giris_zorunlu
def api_odev_mufredat_ozet():
    from database import odev_mufredat_ozeti

    return jsonify(odev_mufredat_ozeti(session["ogretmen_id"]))


@app.route("/manifest.json")
def manifest():
    return app.send_static_file("manifest.json")


@app.route("/sw.js")
def service_worker():
    resp = make_response(app.send_static_file("sw.js"))
    resp.headers["Service-Worker-Allowed"] = "/"
    return resp


# ══════════════════════════════════════════════════════════════════════════
# Tik Islemleri
# ══════════════════════════════════════════════════════════════════════════

@app.route("/tik/<int:ogrenci_id>", methods=["POST"])
@giris_zorunlu
def tik_at(ogrenci_id):
    oid = session["ogretmen_id"]
    if not _ogretmen_ogrencisine_erisebilir(oid, ogrenci_id):
        abort(403)
    kriter   = request.form.get("kriter", "Diger")
    sinif_id = request.form.get("sinif_id", "")

    onceki = ogrenci_tik_sayisi(ogrenci_id)
    if onceki >= OLUMSUZ_TIK_LIMIT:
        mesaj = f"Olumsuz tik limiti doldu ({OLUMSUZ_TIK_LIMIT}/12). Idari islem baslatilmali."
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            d = _durum(onceki)
            return jsonify({
                "ok": False,
                "sebep": mesaj,
                "tik_sayisi": onceki,
                "durum": d["kod"],
                "emoji": d["emoji"],
                "etiket": d["etiket"],
                "sinirda": True,
                "idari_islem": True,
            }), 400
        flash(mesaj, "error")
        return redirect(url_for("dashboard", sinif=sinif_id))

    yeni  = tik_ekle(ogrenci_id, oid, kriter)
    d     = _durum(yeni)

    yeni_seviye = None
    for esik, _, _, etiket in TIK_SEVIYELERI:
        if esik > 0 and onceki < esik <= yeni:
            yeni_seviye = etiket
            break
    idari_islem = yeni >= OLUMSUZ_TIK_LIMIT
    if idari_islem and onceki < OLUMSUZ_TIK_LIMIT:
        try:
            denetim_kaydet(
                "idari_islem_baslat",
                f"{kriter} ile {OLUMSUZ_TIK_LIMIT}/12 olumsuz tik limitine ulasildi.",
                ogretmen_id=oid,
                ogrenci_id=ogrenci_id,
            )
        except Exception:
            pass

    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return jsonify({
            "ok": True,
            "tik_sayisi":  yeni,
            "durum":       d["kod"],
            "emoji":       d["emoji"],
            "etiket":      d["etiket"],
            "yeni_seviye": yeni_seviye,
            "uyari":       idari_islem and onceki < OLUMSUZ_TIK_LIMIT,
            "idari_islem": idari_islem,
            "limit":       OLUMSUZ_TIK_LIMIT,
        })
    return redirect(url_for("dashboard", sinif=sinif_id))


@app.route("/gecmis/<int:ogrenci_id>")
def gecmis(ogrenci_id):
    if "ogretmen_id" in session:
        if not _ogretmen_ogrencisine_erisebilir(session["ogretmen_id"], ogrenci_id):
            return jsonify({"ok": False, "sebep": "Yetkisiz"}), 403
        return jsonify(ogrenci_tik_gecmisi(ogrenci_id))
    oid = _kendi_ogrenci_id_veli_veya_ogrenci()
    if oid == ogrenci_id:
        return jsonify(ogrenci_tik_gecmisi(ogrenci_id))
    if not session.get("ogretmen_id") and not session.get("ogrenci_giris") and not session.get("veli_ogrenci_id"):
        return redirect(url_for("login"))
    return jsonify({"ok": False, "sebep": "Yetkisiz"}), 403


# ══════════════════════════════════════════════════════════════════════════
# Sifirlama
# ══════════════════════════════════════════════════════════════════════════

@app.route("/sifirla/ogrenci/<int:ogrenci_id>", methods=["POST"])
@giris_zorunlu
def sifirla_ogrenci(ogrenci_id):
    if not _ogretmen_ogrencisine_erisebilir(session["ogretmen_id"], ogrenci_id):
        abort(403)
    silinen = ogretmenin_ogrenci_tiklerini_sifirla(ogrenci_id, session["ogretmen_id"])
    sinif_id = request.form.get("sinif_id", "")
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return jsonify({"ok": True, "silinen": silinen})
    return redirect(url_for("dashboard", sinif=sinif_id))


@app.route("/sifirla/sinif/<int:sinif_id>", methods=["POST"])
@giris_zorunlu
def sifirla_sinif(sinif_id):
    if not _ogretmen_sinifinda_mi(session["ogretmen_id"], sinif_id):
        abort(403)
    if request.form.get("parola") == SIFIR_PAROLA:
        ogretmenin_sinif_tiklerini_sifirla(sinif_id, session["ogretmen_id"])
    return redirect(url_for("dashboard", sinif=sinif_id))


@app.route("/sifirla/hepsi", methods=["POST"])
@giris_zorunlu
def sifirla_hepsi():
    if not _toplu_sifirlamaya_izinli_mi(session["ogretmen_id"]):
        abort(403)
    if request.form.get("parola") == SIFIR_PAROLA:
        tum_tikleri_sifirla()
    return redirect(url_for("dashboard"))


# ══════════════════════════════════════════════════════════════════════════
# Olumlu Davranis + Super Lig
# ══════════════════════════════════════════════════════════════════════════

@app.route("/olumlu/<int:sinif_id>", methods=["POST"])
@giris_zorunlu
def olumlu_ekle(sinif_id):
    kriter = request.form.get("kriter", "").strip()
    oid = session["ogretmen_id"]
    ogrenci_id = request.form.get("ogrenci_id", type=int)
    if not kriter or not ogrenci_id:
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({"ok": False, "sebep": "Eksik veri"}), 400
        return redirect(url_for("dashboard", sinif=sinif_id))
    if not _ogretmen_sinifinda_mi(oid, sinif_id):
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({"ok": False, "sebep": "Bu sınıf için yetkiniz yok."}), 403
        abort(403)
    if not _ogretmen_ogrencisine_erisebilir(oid, ogrenci_id):
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({"ok": False, "sebep": "Bu öğrenci için yetkiniz yok."}), 403
        abort(403)
    og = _ogrenci_bul(ogrenci_id)
    if not og or int(og["sinif_id"]) != int(sinif_id):
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({"ok": False, "sebep": "Öğrenci bu sınıfa ait değil."}), 403
        abort(403)
    sonuc = olumlu_tik_ekle(ogrenci_id, sinif_id, oid, kriter)
    if not sonuc.get("ok", True):
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify(sonuc), 400
        flash(sonuc.get("sebep") or "Olumlu tik eklenemedi.", "error")
        return redirect(url_for("dashboard", sinif=sinif_id))

    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return jsonify({"ok": True, "limit": OLUMLU_TIK_LIMIT, **sonuc})
    return redirect(url_for("dashboard", sinif=sinif_id))


@app.route("/api/ogretmen/ogrenci/<int:ogrenci_id>")
@giris_zorunlu
def api_ogretmen_ogrenci_ozet(ogrenci_id: int):
    if not _ogretmen_ogrencisine_erisebilir(session["ogretmen_id"], ogrenci_id):
        return jsonify({"ok": False, "sebep": "Yetkisiz"}), 403
    oz = _ogrenci_ozeti(ogrenci_id)
    if not oz:
        return jsonify({"ok": False}), 404
    gel = oz.get("gelisim") or {}
    xp = int((gel.get("puan") or {}).get("xp") or 0)
    return jsonify({
        "ok": True,
        "ogrenci": oz["ogrenci"],
        "xp": xp,
        "avatar_gelisim": gel.get("avatar") or {},
        "rozetler": oz["rozetler"],
        "istatistik": oz["istatistik"],
        "sezon": oz.get("sezon"),
        "disiplin_sira": oz.get("disiplin_sira"),
        "gelisim": gel,
    })


@app.route("/api/lig")
@giris_zorunlu
def api_lig():
    return jsonify(lig_siralama())


@app.route("/api/lig/sifirla", methods=["POST"])
@giris_zorunlu
def api_lig_sifirla():
    if not _toplu_sifirlamaya_izinli_mi(session["ogretmen_id"]):
        return jsonify({"ok": False, "sebep": "Yetkisiz"}), 403
    if request.form.get("parola") == SIFIR_PAROLA:
        lig_manuel_sifirla()
        return jsonify({"ok": True})
    return jsonify({"ok": False, "hata": "Yanlis parola"}), 403


@app.route("/api/olumlu/gecmis/<int:sinif_id>")
@giris_zorunlu
def api_olumlu_gecmis(sinif_id):
    return jsonify(sinif_olumlu_gecmis(sinif_id))


# ══════════════════════════════════════════════════════════════════════════
# Yayin Modu
# ══════════════════════════════════════════════════════════════════════════

@app.route("/yayin")
@app.route("/yayin/<int:sinif_id>")
@giris_zorunlu
def yayin(sinif_id=None):
    siniflar = _tum_siniflar()
    yayin_veri = _yayin_verisi_hazirla()
    gl = gelisim_ligi()
    ticker_lider = [{"sinif_adi": x["sinif_adi"], "puan": x.get("xp", 0)} for x in gl[:8]]
    return render_template(
        "yayin.html",
        sinif_sayisi=len(siniflar),
        yayin_veri=yayin_veri,
        liderler=ticker_lider,
        rozetler=son_rozetler(8),
        sezon=sezon_siralama()[:5],
    )


# ══════════════════════════════════════════════════════════════════════════
# Ogrenci Goruntulemesi (sifreli, salt okunur)
# ══════════════════════════════════════════════════════════════════════════

@app.route("/ogrenci")
@ogrenci_giris_zorunlu
def ogrenci_gorunum():
    ogrenci_id = session.get("ogrenci_id")
    if not ogrenci_id:
        return redirect(url_for("ogrenci_giris", next=url_for("ogrenci_gorunum")))
    o = _ogrenci_bul(int(ogrenci_id))
    if not o:
        session.pop("ogrenci_id", None)
        session.pop("ogrenci_giris", None)
        return redirect(url_for("ogrenci_giris"))
    gecmis = ogrenci_tik_gecmisi(int(ogrenci_id))
    odevler = ogrenci_odevleri(int(ogrenci_id), 40)
    return render_template(
        "tik_gecmisi.html",
        ogrenci=o,
        gecmis=gecmis,
        avatar=_avatar(o),
        veli_modu=False,
        odevler=odevler,
    )


@app.route("/api/kendi-tik-gecmisi")
def api_kendi_tik_gecmisi():
    oid = _kendi_ogrenci_id_veli_veya_ogrenci()
    if not oid:
        return jsonify({"ok": False, "sebep": "Giris gerekli"}), 401
    o = _ogrenci_bul(int(oid))
    if not o:
        return jsonify({"ok": False, "sebep": "Ogrenci bulunamadi"}), 404
    return jsonify({
        "ok": True,
        "ogrenci": o,
        "gecmis": ogrenci_tik_gecmisi(int(oid)),
        "odevler": ogrenci_odevleri(int(oid), 40),
    })


@app.route("/api/ogrenci/veri")
@ogrenci_giris_zorunlu
def api_ogrenci_veri():
    """Eski uç nokta; artık yalnızca oturumdaki öğrencinin tik verisini döndürür."""
    return api_kendi_tik_gecmisi()


@app.route("/api/yayin/ogrenciler")
@giris_zorunlu
def api_yayin_ogrenciler():
    return jsonify(_yayin_verisi_hazirla())


def _yayin_verisi_hazirla():
    ogrenciler = _ogrencilere_durum_ekle(tum_okul_ogrencileri())
    ids = [o["id"] for o in ogrenciler]
    rozet_haritasi = ogrenci_rozetleri_yayin_map(ids, limit=8)
    env_haritasi = ogrenci_aktif_envanter_map(ids)
    olumlu_h = ogrenci_olumlu_tik_sayilari(ids)
    for o in ogrenciler:
        o["risk_yuzde"] = min(100, o["tik_sayisi"] * 8)
        o["rozetler"] = rozet_haritasi.get(o["id"], [])
        o["aktif_envanter"] = env_haritasi.get(o["id"])
        o["olumlu_tik"] = olumlu_h.get(o["id"], 0)
    en_cok_olumsuz = sorted(
        ogrenciler,
        key=lambda o: (-int(o.get("tik_sayisi") or 0), o.get("sinif_adi") or "", o.get("ad_soyad") or ""),
    )
    en_cok_olumlu = sorted(
        ogrenciler,
        key=lambda o: (-int(o.get("olumlu_tik") or 0), o.get("sinif_adi") or "", o.get("ad_soyad") or ""),
    )
    ogrenciler = en_cok_olumsuz
    siniflar = {}
    for o in ogrenciler:
        s = siniflar.setdefault(o["sinif_adi"], {"sinif_adi": o["sinif_adi"], "ogrenci": 0, "tik": 0, "temiz": 0, "idari": 0})
        s["ogrenci"] += 1
        s["tik"] += o["tik_sayisi"]
        s["temiz"] += 1 if o["tik_sayisi"] == 0 else 0
        s["idari"] += 1 if o["tik_sayisi"] >= OLUMSUZ_TIK_LIMIT else 0
    odevler = []
    for s in _tum_siniflar():
        for od in sinif_odevleri(s["id"], 6):
            odevler.append({**od, "sinif_adi": s["sinif_adi"]})
    odevler.sort(key=lambda od: (od.get("son_tarih", ""), od.get("id", 0)), reverse=True)
    return {
        "ogrenciler": ogrenciler,
        "en_cok_olumsuz": en_cok_olumsuz,
        "en_cok_olumlu": en_cok_olumlu,
        "siniflar": sorted(siniflar.values(), key=lambda s: (-s["tik"], s["sinif_adi"])),
        "odevler": odevler[:16],
        "ozet": {
            "ogrenci": len(ogrenciler),
            "sinif": len(siniflar),
            "tik": sum(o["tik_sayisi"] for o in ogrenciler),
            "temiz": sum(1 for o in ogrenciler if o["tik_sayisi"] == 0),
            "uyari": sum(1 for o in ogrenciler if 1 <= o["tik_sayisi"] <= 2),
            "idari": sum(1 for o in ogrenciler if o["tik_sayisi"] >= OLUMSUZ_TIK_LIMIT),
            "veli": sum(1 for o in ogrenciler if 6 <= o["tik_sayisi"] <= 8),
            "tutanak": sum(1 for o in ogrenciler if 9 <= o["tik_sayisi"] <= 11),
            "disiplin": sum(1 for o in ogrenciler if o["tik_sayisi"] >= 12),
        }
    }


@app.route("/api/yayin/<int:sinif_id>")
@giris_zorunlu
def api_yayin_sinif(sinif_id):
    return jsonify(_ogrencilere_durum_ekle(sinif_ogrencileri(sinif_id)))


@app.route("/api/yayin/lig")
@giris_zorunlu
def api_yayin_lig():
    return jsonify(gelisim_ligi())


# ══════════════════════════════════════════════════════════════════════════
# Excel Raporu
# ══════════════════════════════════════════════════════════════════════════

@app.route("/rapor/excel")
@giris_zorunlu
def rapor_excel():
    if not OPENPYXL_OK:
        return "openpyxl kurulu degil", 500

    ogretmen_id  = session["ogretmen_id"]
    ogretmen_adi = session["ogretmen_adi"]
    siniflar     = ogretmen_siniflari(ogretmen_id)
    sinif_id_str = request.args.get("sinif_id")
    yalnizca_id  = int(sinif_id_str) if sinif_id_str else None

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    excel_raporu_olustur(tmp_path, ogretmen_adi, siniflar, yalnizca_id)
    tarih     = datetime.now().strftime("%Y%m%d_%H%M")
    dosya_adi = f"DisiplinRaporu_{tarih}.xlsx"

    return send_file(tmp_path, as_attachment=True, download_name=dosya_adi,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/rapor/analiz-pdf")
@giris_zorunlu
def rapor_analiz_pdf():
    if not PDF_OK:
        return (
            "PDF icin reportlab gerekli: pip install reportlab",
            500,
        )
    ogretmen_id = session["ogretmen_id"]
    ogretmen_adi = session.get("ogretmen_adi", "")
    siniflar = ogretmen_siniflari(ogretmen_id)
    sinif_id_str = request.args.get("sinif_id")
    yalnizca_id = int(sinif_id_str) if sinif_id_str else None

    snapshot = derle_analiz_snapshot(siniflar, yalnizca_id)
    try:
        pdf_bytes = pdf_analiz_uret_bytes(snapshot, ogretmen_adi)
    except ImportError as e:
        return str(e), 500
    json_txt = json.dumps(snapshot, ensure_ascii=False)
    kapsam = (snapshot.get("meta") or {}).get("kapsam_metin", "")[:500]
    tarih = datetime.now().strftime("%Y%m%d_%H%M")
    dosya_adi = f"DisiplinAnaliz_{tarih}.pdf"
    rapor_arsiv_kaydet(
        ogretmen_id,
        ogretmen_adi,
        kapsam or "Analiz",
        yalnizca_id,
        json_txt,
        pdf_bytes,
        dosya_adi,
    )
    return send_file(
        BytesIO(pdf_bytes),
        as_attachment=True,
        download_name=dosya_adi,
        mimetype="application/pdf",
    )


@app.route("/analiz")
@giris_zorunlu
def analiz_merkezi():
    ogretmen_id = session["ogretmen_id"]
    siniflar = ogretmen_siniflari(ogretmen_id)
    secili_sinif_id = request.args.get("sinif_id", type=int)
    if secili_sinif_id and secili_sinif_id not in {s["id"] for s in siniflar}:
        secili_sinif_id = None
    admin_mi = _toplu_sifirlamaya_izinli_mi(ogretmen_id)
    return render_template(
        "analiz_merkezi.html",
        siniflar=siniflar,
        secili_sinif_id=secili_sinif_id,
        pdf_ok=PDF_OK,
        arsiv=rapor_arsiv_listesi(ogretmen_id, 12),
        sistem_yedekleri=sistem_yedek_listesi(10) if admin_mi else [],
        admin_mi=admin_mi,
        ogretmen_adi=session.get("ogretmen_adi", ""),
    )


@app.route("/rapor/arsiv")
@giris_zorunlu
def rapor_arsiv_sayfa():
    oid = session["ogretmen_id"]
    kayitlar = rapor_arsiv_listesi(oid, 60)
    yedek_gruplari = rapor_arsiv_yedek_gruplari(oid)
    return render_template(
        "rapor_arsiv.html",
        kayitlar=kayitlar,
        yedek_gruplari=yedek_gruplari,
    )


@app.route("/rapor/arsiv/sifirla", methods=["POST"])
@giris_zorunlu
def rapor_arsiv_sifirla():
    sonuc = rapor_arsiv_tumunu_yedekle_ve_sil(session["ogretmen_id"])
    if sonuc.get("tasinan", 0) == 0:
        flash("Aktif PDF arşivinde silinecek kayıt yok.", "info")
    else:
        flash(
            f"{sonuc['tasinan']} analiz raporu güvenli yedeğe alındı; liste temizlendi. "
            "Aşağıdan «Geri yükle» ile istediğiniz silme anına dönebilirsiniz.",
            "success",
        )
    return redirect(url_for("rapor_arsiv_sayfa"))


@app.route("/rapor/arsiv/yedek/<path:grup_id>/geri-yukle", methods=["POST"])
@giris_zorunlu
def rapor_arsiv_yedek_geri_yukle(grup_id: str):
    sonuc = rapor_arsiv_grubu_geri_yukle(session["ogretmen_id"], grup_id)
    if sonuc.get("ok"):
        flash(f"{sonuc['geri_yuklenen']} rapor tekrar arşive eklendi.", "success")
    else:
        flash(sonuc.get("sebep", "Geri yükleme yapılamadı."), "error")
    return redirect(url_for("rapor_arsiv_sayfa"))


@app.route("/rapor/arsiv/<int:arsiv_id>/indir")
@giris_zorunlu
def rapor_arsiv_indir(arsiv_id: int):
    row = rapor_arsiv_pdf_oku(arsiv_id, session["ogretmen_id"])
    if not row or not row.get("pdf_blob"):
        abort(404)
    return send_file(
        BytesIO(row["pdf_blob"]),
        as_attachment=True,
        download_name=row.get("dosya_adi") or f"rapor_{arsiv_id}.pdf",
        mimetype="application/pdf",
    )


@app.route("/rapor/excel-detayli")
@giris_zorunlu
def rapor_excel_detayli():
    if not OPENPYXL_OK:
        return "openpyxl kurulu degil", 500

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import BarChart, Reference

    wb = Workbook()
    ws = wb.active
    ws.title = "Okul Detay Raporu"
    baslik = ["Sıra", "Sınıf", "No", "Ad Soyad", "Tik", "Durum", "Son Tik", "Ödev", "XP", "Risk"]
    ws.append(baslik)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1A3A6B")
        cell.alignment = Alignment(horizontal="center")

    ogrenciler = _ogrencilere_durum_ekle(tum_okul_ogrencileri())
    ogrenciler.sort(key=lambda o: (-o["tik_sayisi"], o["sinif_adi"], o["ad_soyad"]))
    for idx, o in enumerate(ogrenciler, 1):
        odevler = ogrenci_odevleri(o["id"], 50)
        tamam = sum(1 for od in odevler if od.get("tamamlandi"))
        odev_orani = round(tamam * 100 / len(odevler)) if odevler else 0
        gel = gelisim_ozeti(o["id"])
        risk = min(100, o["tik_sayisi"] * 8 + (15 if odev_orani < 50 and odevler else 0))
        gecmis = ogrenci_tik_gecmisi(o["id"])
        ws.append([
            idx, o["sinif_adi"], o["ogr_no"], o["ad_soyad"], o["tik_sayisi"],
            o["etiket"] or "Temiz", gecmis[0]["tarih"] if gecmis else "-",
            f"%{odev_orani}", gel["puan"]["xp"], risk,
        ])

    ws2 = wb.create_sheet("Sınıf Özeti")
    ws2.append(["Sınıf", "Öğrenci", "Toplam Tik", "Temiz", "İdari"])
    sinif_ozet = {}
    for o in ogrenciler:
        s = sinif_ozet.setdefault(o["sinif_adi"], {"ogrenci": 0, "tik": 0, "temiz": 0, "idari": 0})
        s["ogrenci"] += 1
        s["tik"] += o["tik_sayisi"]
        s["temiz"] += 1 if o["tik_sayisi"] == 0 else 0
        s["idari"] += 1 if o["tik_sayisi"] >= OLUMSUZ_TIK_LIMIT else 0
    for ad, s in sorted(sinif_ozet.items()):
        ws2.append([ad, s["ogrenci"], s["tik"], s["temiz"], s["idari"]])
    chart = BarChart()
    chart.title = "Sınıf Bazlı Toplam Tik"
    data = Reference(ws2, min_col=3, min_row=1, max_row=ws2.max_row)
    cats = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws2.add_chart(chart, "G2")

    for sheet in wb.worksheets:
        for col in sheet.columns:
            sheet.column_dimensions[col[0].column_letter].width = min(35, max(12, max(len(str(c.value or "")) for c in col) + 2))

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    wb.save(tmp_path)
    tarih = datetime.now().strftime("%Y%m%d_%H%M")
    return send_file(tmp_path, as_attachment=True,
                     download_name=f"DetayliOkulRaporu_{tarih}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ══════════════════════════════════════════════════════════════════════════
# Süper Lig — maçlar, tablo, kart, VAR
# ══════════════════════════════════════════════════════════════════════════

@app.route("/lig")
@giris_zorunlu
def lig():
    maclar = bugun_maclar()
    tablo = lig_puan_tablosu()
    return render_template(
        "lig.html",
        maclar=maclar,
        tablo=tablo,
        ogretmen_id=session["ogretmen_id"],
        var_ogretmenler=_var_hakem_idleri(),
        kart_nedenleri=LIG_KART_NEDENLERI,
    )


@app.route("/lig/olustur", methods=["POST"])
@giris_zorunlu
def lig_olustur():
    gunluk_mac_olustur()
    return redirect(url_for("lig"))


@app.route("/api/lig/maclar")
@giris_zorunlu
def api_lig_maclar():
    return jsonify(bugun_maclar())


@app.route("/api/lig/tablo")
@giris_zorunlu
def api_lig_tablo():
    return jsonify(lig_puan_tablosu())


@app.route("/lig/mac/<int:mac_id>")
@giris_zorunlu
def lig_mac_detay(mac_id):
    return redirect(url_for("lig"))


@app.route("/lig/mac/<int:mac_id>/sonuc", methods=["POST"])
@giris_zorunlu
def lig_mac_sonuc(mac_id):
    s1t = request.form.get("s1_tamamlayan", type=int)
    s2t = request.form.get("s2_tamamlayan", type=int)
    s1top = request.form.get("s1_toplam", type=int)
    s2top = request.form.get("s2_toplam", type=int)
    sonuc = mac_sonucu_gir(
        mac_id,
        s1t if s1t is not None else 0,
        max(1, s1top or 1),
        s2t if s2t is not None else 0,
        max(1, s2top or 1),
    )
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return jsonify(sonuc)
    return redirect(url_for("lig"))


@app.route("/lig/mac/<int:mac_id>/oyla", methods=["POST"])
@giris_zorunlu
def lig_mac_oyla(mac_id):
    sid = request.form.get("sinif_id", type=int)
    if not sid:
        return jsonify({"durum": "hata", "sebep": "sinif_yok"}), 400
    sonuc = mac_oy_ver(mac_id, session["ogretmen_id"], sid)
    return jsonify(sonuc)


@app.route("/lig/sifirla", methods=["POST"])
@giris_zorunlu
def lig_sifirla_mac():
    if request.form.get("parola") != SIFIR_PAROLA:
        flash("Yanlış parola.", "error")
        return redirect(url_for("lig"))
    lig_mac_tablo_sifirla()
    flash("Sezon maç ve tablo sıfırlandı.", "success")
    return redirect(url_for("lig"))


@app.route("/api/kadro/<int:sinif_id>")
@giris_zorunlu
def api_kadro(sinif_id):
    return jsonify([])


@app.route("/api/ogrenci/kadro/<int:sinif_id>")
@ogrenci_giris_zorunlu
def api_ogrenci_kadro(sinif_id):
    return jsonify([])


@app.route("/api/ogrenci/ittifaklar")
@ogrenci_giris_zorunlu
def api_ogrenci_ittifaklar():
    return jsonify(aktif_ittifaklar())


@app.route("/api/kadro/<int:sinif_id>/yenile", methods=["POST"])
@giris_zorunlu
def api_kadro_yenile(sinif_id):
    return jsonify({"ok": False, "sebep": "Kadro sistemi kaldirildi"})


@app.route("/api/sinif/<int:sinif_id>/ogrenciler")
@giris_zorunlu
def api_sinif_ogrencileri_json(sinif_id):
    return jsonify(sinif_ogrencileri(sinif_id))


@app.route("/lig/mac/<int:mac_id>/kart", methods=["POST"])
@giris_zorunlu
def lig_mac_kart(mac_id):
    ogrenci_id = request.form.get("ogrenci_id", type=int)
    sinif_id = request.form.get("sinif_id", type=int)
    kart_turu = (request.form.get("kart_turu") or "").strip()
    neden = (request.form.get("neden") or "Kural ihlali").strip()[:120]
    if not ogrenci_id or not sinif_id or kart_turu not in ("sari", "kirmizi"):
        return jsonify({"ok": False, "hata": "Eksik veya geçersiz veri"}), 400
    if not _ogretmen_ogrencisine_erisebilir(session["ogretmen_id"], ogrenci_id):
        return jsonify({"ok": False, "hata": "Yetkisiz"}), 403
    sonuc = kart_ver(mac_id, ogrenci_id, sinif_id, session["ogretmen_id"], kart_turu, neden)
    return jsonify({"ok": True, **sonuc})


@app.route("/api/lig/mac/<int:mac_id>/kartlar")
@giris_zorunlu
def api_mac_kartlari(mac_id):
    return jsonify(mac_kartlari(mac_id))


@app.route("/api/lig/maclar_ve_tablo")
@giris_zorunlu
def api_lig_maclar_ve_tablo():
    return jsonify({
        "maclar": bugun_maclar(),
        "tablo": lig_puan_tablosu(),
        "kadrolar": {},
    })


# ══════════════════════════════════════════════════════════════════════════
# Gamification Rotalar
# ══════════════════════════════════════════════════════════════════════════

@app.route("/oduller")
@giris_zorunlu
def oduller():
    ogretmen_adi = session["ogretmen_adi"]
    return render_template("oduller.html",
        ogretmen_adi    = ogretmen_adi,
        ogretmen_id     = session["ogretmen_id"],
        mufettis_yetkili = True,
        sezon           = sezon_siralama(),
        seviyeleri      = tum_sinif_seviyeleri(),
        rozetler        = son_rozetler(30),
        seri            = tum_seri_tablosu(),
        gorev           = bugun_gorev(),
        mufettis        = bugun_mufettis(),
        ittifaklar      = aktif_ittifaklar(),
        bekleyen_talepler = bekleyen_ogrenci_talepleri(),
        alkislar        = son_alkislar(15),
        rozet_tanimi    = ROZET_TANIMI,
        sans_secenekler = SANS_CARKI_SEENEKLERI,
    )


@app.route("/api/oduller/veri")
@giris_zorunlu
def api_oduller_veri():
    return jsonify({
        "sezon":    sezon_siralama(),
        "seviye":   tum_sinif_seviyeleri(),
        "rozetler": son_rozetler(20),
        "seri":     tum_seri_tablosu(),
        "gorev":    bugun_gorev(),
        "alkislar": son_alkislar(10),
        "ittifaklar": aktif_ittifaklar(),
    })


@app.route("/api/gorev/tamamla", methods=["POST"])
@giris_zorunlu
def api_gorev_tamamla():
    gorev_id = int(request.form["gorev_id"])
    sinif_id = int(request.form["sinif_id"])
    return jsonify(gorev_tamamla(gorev_id, sinif_id))


@app.route("/api/mufettis/degerlendir", methods=["POST"])
@giris_zorunlu
def api_mufettis_degerlendir():
    mufettis_id = int(request.form["mufettis_id"])
    sonuc       = request.form["sonuc"]
    if sonuc not in ("iyi", "kotu"):
        return jsonify({"ok": False}), 400
    return jsonify(mufettis_degerlendir(mufettis_id, sonuc))


@app.route("/api/alkis/ver", methods=["POST"])
@giris_zorunlu
def api_alkis_ver():
    ogrenci_id  = int(request.form["ogrenci_id"])
    sinif_id    = int(request.form["sinif_id"])
    mesaj       = request.form.get("mesaj", "Harika is!")
    ogretmen_id = session["ogretmen_id"]
    return jsonify(alkis_ver(ogrenci_id, sinif_id, ogretmen_id, mesaj))


@app.route("/api/ittifak/olustur", methods=["POST"])
@giris_zorunlu
def api_ittifak_olustur():
    sinif1_id = int(request.form["sinif1_id"])
    sinif2_id = int(request.form["sinif2_id"])
    return jsonify(ittifak_olustur(sinif1_id, sinif2_id))


@app.route("/api/ittifak/tamamla", methods=["POST"])
@giris_zorunlu
def api_ittifak_tamamla():
    ittifak_id = int(request.form["ittifak_id"])
    return jsonify(ittifak_tamamla(ittifak_id))


@app.route("/api/ittifak/onayla", methods=["POST"])
@giris_zorunlu
def api_ittifak_onayla():
    ittifak_id = int(request.form["ittifak_id"])
    return jsonify(ittifak_onayla(ittifak_id))


@app.route("/api/ittifak/reddet", methods=["POST"])
@giris_zorunlu
def api_ittifak_reddet():
    ittifak_id = int(request.form["ittifak_id"])
    return jsonify(ittifak_reddet(ittifak_id))


@app.route("/api/ogrenci/ittifak/talep", methods=["POST"])
@ogrenci_giris_zorunlu
def api_ogrenci_ittifak_talep():
    try:
        sinif1_id = int(request.form["sinif1_id"])
        sinif2_id = int(request.form["sinif2_id"])
        seans     = request.form.get("seans", "sabah")
    except (KeyError, ValueError):
        return jsonify({"ok": False, "sebep": "Eksik bilgi"}), 400
    if sinif1_id == sinif2_id:
        return jsonify({"ok": False, "sebep": "Ayni sinif secilemez"})
    return jsonify(ittifak_ogrenci_talebi(sinif1_id, sinif2_id, seans))


@app.route("/api/mufettis/belirle", methods=["POST"])
@giris_zorunlu
def api_mufettis_belirle():
    ogrenci_id = int(request.form["ogrenci_id"])
    return jsonify(mufettis_belirle(ogrenci_id))


@app.route("/api/sans-carki/cevir", methods=["POST"])
@giris_zorunlu
def api_sans_carki():
    return jsonify({"ok": False, "sebep": "Sans carki mac sistemine bagliydi; maçlar kapalı."}), 410


@app.route("/api/tik/dondur", methods=["POST"])
@giris_zorunlu
def api_tik_dondur():
    ogrenci_id  = int(request.form["ogrenci_id"])
    ogretmen_id = session["ogretmen_id"]
    if not _ogretmen_ogrencisine_erisebilir(ogretmen_id, ogrenci_id):
        return jsonify({"ok": False, "sebep": "Yetkisiz"}), 403
    return jsonify(tik_dondur(ogrenci_id, ogretmen_id))


# ══════════════════════════════════════════════════════════════════════════
# Quiz — Bilgi Yarisması (sifresiz)
# ══════════════════════════════════════════════════════════════════════════

QUIZ_DERSLER = {
    5: ["Türkçe", "Matematik", "Fen Bilimleri", "Sosyal Bilgiler", "İngilizce", "Din Kültürü"],
    6: ["Türkçe", "Matematik", "Fen Bilimleri", "Sosyal Bilgiler", "İngilizce", "Din Kültürü"],
    7: ["Türkçe", "Matematik", "Fen Bilimleri", "Sosyal Bilgiler", "İngilizce", "Din Kültürü"],
    8: ["Türkçe", "Matematik", "Fen Bilimleri", "T.C. İnkılap Tarihi", "İngilizce", "Din Kültürü"],
}

DERS_EMOJI = {
    "Türkçe": "📚", "Matematik": "🔢", "Fen Bilimleri": "🔬",
    "Sosyal Bilgiler": "🌍", "T.C. İnkılap Tarihi": "🏛️",
    "İngilizce": "🇬🇧", "Din Kültürü": "🕌",
}


@app.route("/quiz/<int:sinif_id>")
def quiz(sinif_id):
    quiz_sorulari_yukle()
    from database import _conn as _db_conn
    con = _db_conn()
    row = con.execute("SELECT sinif_adi FROM siniflar WHERE id=?", (sinif_id,)).fetchone()
    con.close()
    sinif_adi = row["sinif_adi"] if row else f"Sinif {sinif_id}"
    bugun_yapilan = quiz_gunluk_dersleri(sinif_id)
    istatistik = quiz_sinif_istatistik(sinif_id)
    return render_template("quiz.html",
        sinif_id=sinif_id, sinif_adi=sinif_adi,
        quiz_dersler=QUIZ_DERSLER,
        ders_emoji=DERS_EMOJI,
        bugun_yapilan=bugun_yapilan,
        istatistik=istatistik,
    )


@app.route("/api/quiz/sorular")
def api_quiz_sorular():
    sinif_seviyesi = request.args.get("sinif_seviyesi", type=int)
    ders = request.args.get("ders", "")
    if not sinif_seviyesi or not ders:
        return jsonify({"ok": False, "sebep": "Eksik parametre"}), 400
    sorular = quiz_sorular_getir(sinif_seviyesi, ders)
    for s in sorular:
        s.pop("dogru_cevap", None)
    return jsonify({"ok": True, "sorular": sorular})


@app.route("/api/quiz/cevapla", methods=["POST"])
def api_quiz_cevapla():
    veri = request.json or {}
    sinif_id = veri.get("sinif_id")
    sinif_seviyesi = veri.get("sinif_seviyesi")
    ders = veri.get("ders")
    cevaplar = veri.get("cevaplar", {})
    if not (sinif_id and sinif_seviyesi and ders and cevaplar):
        return jsonify({"ok": False, "sebep": "Eksik veri"}), 400
    from database import _conn as _dbc
    from database import _quiz_init as _qi
    con = _dbc()
    _qi(con)
    dogru = yanlis = 0
    sonuclar = {}
    for soru_id_str, verilen in cevaplar.items():
        row = con.execute("SELECT dogru_cevap FROM quiz_sorulari WHERE id=?",
                          (int(soru_id_str),)).fetchone()
        if row:
            gercek = row["dogru_cevap"]
            ok = verilen.upper() == gercek.upper()
            sonuclar[soru_id_str] = {"dogru": ok, "gercek": gercek}
            if ok: dogru += 1
            else:  yanlis += 1
    con.close()
    sonuc = quiz_sonuc_kaydet(sinif_id, ders, sinif_seviyesi, dogru, yanlis)
    sonuc["dogru"] = dogru
    sonuc["yanlis"] = yanlis
    sonuc["sonuclar"] = sonuclar
    return jsonify(sonuc)


@app.route("/api/quiz/istatistik/<int:sinif_id>")
def api_quiz_istatistik(sinif_id):
    return jsonify(quiz_sinif_istatistik(sinif_id))


@app.route("/taktik/<int:sinif_id>")
@giris_zorunlu
def taktik(sinif_id):
    ogretmen_id = int(session["ogretmen_id"])
    if not _ogretmen_sinifinda_mi(ogretmen_id, sinif_id) and not _toplu_sifirlamaya_izinli_mi(ogretmen_id):
        abort(403)
    sinif = next((s for s in _tum_siniflar() if int(s["id"]) == int(sinif_id)), None)
    if not sinif:
        abort(404)
    return render_template(
        "taktik.html",
        sinif_id=sinif_id,
        sinif_adi=sinif["sinif_adi"],
        kadro=sinif_ogrencileri(sinif_id),
        kayit=taktik_yukle(sinif_id),
        mevkiler_json=_mevkiler_json(),
        ana_menu_url=url_for("dashboard", sinif=sinif_id),
        kaydet_url=url_for("api_taktik_kaydet", sinif_id=sinif_id),
    )


@app.route("/api/taktik/<int:sinif_id>")
@giris_zorunlu
def api_taktik_yukle(sinif_id):
    ogretmen_id = int(session["ogretmen_id"])
    if not _ogretmen_sinifinda_mi(ogretmen_id, sinif_id) and not _toplu_sifirlamaya_izinli_mi(ogretmen_id):
        abort(403)
    return jsonify(taktik_yukle(sinif_id))


@app.route("/api/taktik/<int:sinif_id>/kaydet", methods=["POST"])
@giris_zorunlu
def api_taktik_kaydet(sinif_id):
    ogretmen_id = int(session["ogretmen_id"])
    if not _ogretmen_sinifinda_mi(ogretmen_id, sinif_id) and not _toplu_sifirlamaya_izinli_mi(ogretmen_id):
        abort(403)
    veri = request.get_json(silent=True) or {}
    return jsonify(taktik_kaydet(sinif_id, json.dumps(veri, ensure_ascii=False)))


@app.route("/ogrenci-maclari")
@giris_zorunlu
def ogretmen_ogrenci_maclari():
    ogretmen_id = int(session["ogretmen_id"])
    admin_mi = _toplu_sifirlamaya_izinli_mi(ogretmen_id)
    bekleyenler = ogretmen_onay_bekleyen_ogrenci_maclari(ogretmen_id, admin_mi, "onay_bekliyor")
    son_islenenler = ogretmen_onay_bekleyen_ogrenci_maclari(ogretmen_id, admin_mi, None)[:30]
    return render_template(
        "ogrenci_mac_onay.html",
        bekleyenler=bekleyenler,
        son_islenenler=son_islenenler,
    )


@app.route("/ogrenci-maclari/<int:mac_id>/onayla", methods=["POST"])
@giris_zorunlu
def ogretmen_ogrenci_mac_onayla(mac_id):
    ogretmen_id = int(session["ogretmen_id"])
    mac = ogrenci_mac_detay(mac_id)
    if not _ogretmen_ogrenci_macina_erisebilir(ogretmen_id, mac):
        abort(403)
    sonuc = ogrenci_mac_onayla(mac_id, ogretmen_id, True)
    flash(
        "Mac onaylandi; kazanan sinifa +3 lig puani islendi." if sonuc.get("ok") else sonuc.get("sebep", "Mac onaylanamadi."),
        "success" if sonuc.get("ok") else "warning",
    )
    return redirect(url_for("ogretmen_ogrenci_maclari"))


@app.route("/ogrenci-maclari/<int:mac_id>/reddet", methods=["POST"])
@giris_zorunlu
def ogretmen_ogrenci_mac_reddet(mac_id):
    ogretmen_id = int(session["ogretmen_id"])
    mac = ogrenci_mac_detay(mac_id)
    if not _ogretmen_ogrenci_macina_erisebilir(ogretmen_id, mac):
        abort(403)
    sonuc = ogrenci_mac_onayla(mac_id, ogretmen_id, False)
    flash(
        "Mac reddedildi." if sonuc.get("ok") else sonuc.get("sebep", "Mac reddedilemedi."),
        "info" if sonuc.get("ok") else "warning",
    )
    return redirect(url_for("ogretmen_ogrenci_maclari"))


# ══════════════════════════════════════════════════════════════════════════
# Veli özeti, randevu, günlük yansıma, davranış hedefi, ek raporlar, denetim
# ══════════════════════════════════════════════════════════════════════════


@app.route("/veli/ozet")
def veli_ozet_sayfa():
    oid = session.get("veli_ogrenci_id")
    if not oid:
        return redirect(url_for("veli_giris"))
    o = _ogrenci_bul(int(oid))
    if not o:
        session.pop("veli_ogrenci_id", None)
        return redirect(url_for("veli_giris"))
    metrics = veli_ozet_metrikleri(int(oid), 30)
    gecmis_kisa = ogrenci_tik_gecmisi(int(oid))[:20]
    return render_template(
        "veli_ozet.html",
        ogrenci=o,
        metrics=metrics,
        gecmis_kisa=gecmis_kisa,
    )


@app.route("/veli/randevu", methods=["POST"])
def veli_randevu_kaydet():
    oid = session.get("veli_ogrenci_id")
    if not oid:
        return redirect(url_for("veli_giris"))
    o = _ogrenci_bul(int(oid))
    if not o:
        return redirect(url_for("veli_giris"))
    mesaj = request.form.get("mesaj", "")
    randevu_talep_ekle(int(oid), int(o["sinif_id"]), mesaj)
    flash("Görüşme talebiniz kaydedildi. Öğretmeniniz en kısa sürede değerlendirecek.", "success")
    return redirect(url_for("veli_ozet_sayfa"))


@app.route("/ogretmen/randevular")
@giris_zorunlu
def ogretmen_randevular():
    siniflar = ogretmen_siniflari(session["ogretmen_id"])
    ids = [s["id"] for s in siniflar]
    talepler = randevu_listesi_siniflar(ids)
    return render_template(
        "ogretmen_randevular.html",
        talepler=talepler,
        siniflar=siniflar,
    )


@app.route("/ogretmen/randevu/<int:talep_id>/durum", methods=["POST"])
@giris_zorunlu
def ogretmen_randevu_durum(talep_id: int):
    ids = [s["id"] for s in ogretmen_siniflari(session["ogretmen_id"])]
    talep = randevu_talep_by_id(talep_id)
    if not talep or talep.get("sinif_id") not in ids:
        abort(403)
    durum = request.form.get("durum", "gorusuldu")
    randevu_durum_guncelle(talep_id, durum)
    flash("Randevu durumu güncellendi.", "success")
    return redirect(url_for("ogretmen_randevular"))


@app.route("/ogrenci/yansima", methods=["GET", "POST"])
@ogrenci_giris_zorunlu
def ogrenci_yansima():
    oid = int(session["ogrenci_id"])
    if request.method == "POST":
        sonuc = gunluk_yansima_ekle(oid, request.form.get("metin", ""))
        if sonuc.get("ok"):
            flash("Bugünün yansıması gönderildi.", "success")
        else:
            flash(sonuc.get("sebep", "Kaydedilemedi"), "error")
        return redirect(url_for("ogrenci_yansima"))
    gecmis = gunluk_yansima_ogrenci_gecmis(oid, 14)
    return render_template("ogrenci_yansima.html", gecmis=gecmis)


@app.route("/ogretmen/yansimalar", methods=["GET"])
@giris_zorunlu
def ogretmen_yansimalar():
    siniflar = ogretmen_siniflari(session["ogretmen_id"])
    ids = [s["id"] for s in siniflar]
    bekleyen = gunluk_yansima_bekleyen_siniflar(ids)
    return render_template("ogretmen_yansimalar.html", bekleyen=bekleyen)


@app.route("/ogretmen/yansima/<int:yid>/degerlendir", methods=["POST"])
@giris_zorunlu
def ogretmen_yansima_degerlendir_route(yid: int):
    ids = [s["id"] for s in ogretmen_siniflari(session["ogretmen_id"])]
    kayit = gunluk_yansima_by_id(yid)
    if not kayit or kayit.get("ogrenci_sinif_id") not in ids:
        abort(403)
    durum = request.form.get("durum", "onaylandi")
    notu = request.form.get("not", "")
    gunluk_yansima_degerlendir(yid, session["ogretmen_id"], durum, notu)
    flash("Yansıma değerlendirildi.", "success")
    return redirect(url_for("ogretmen_yansimalar"))


@app.route("/api/pozitif-rozet/<int:ogrenci_id>", methods=["POST"])
@giris_zorunlu
def api_pozitif_rozet(ogrenci_id: int):
    if not _ogretmen_ogrencisine_erisebilir(session["ogretmen_id"], ogrenci_id):
        return jsonify({"ok": False, "sebep": "Yetki"}), 403
    og = _ogrenci_bul(ogrenci_id)
    if not og:
        return jsonify({"ok": False}), 404
    yeni = rozet_ver_ogrenci(ogrenci_id, og["sinif_id"], "pozitif_yildiz")
    return jsonify({"ok": True, "yeni": bool(yeni)})


@app.route("/rapor/haftalik")
@giris_zorunlu
def rapor_haftalik():
    siniflar = ogretmen_siniflari(session["ogretmen_id"])
    sid = request.args.get("sinif_id", type=int)
    if siniflar:
        ids = [s["id"] for s in siniflar]
        if sid not in ids:
            sid = siniflar[0]["id"]
    else:
        sid = None
    ozet = haftalik_sinif_ozeti(sid, 7) if sid else {}
    metin_satirlari = []
    if sid and ozet:
        metin_satirlari = [
            f"Haftalık özet ({ozet['baslangic']} — {ozet['bitis']})",
            f"Tik kayıtları: {ozet['tik']}",
            f"Olumlu davranış kayıtları: {ozet['olumlu']}",
        ]
    return render_template(
        "rapor_haftalik.html",
        siniflar=siniflar,
        sinif_id=sid,
        ozet=ozet,
        metin_satirlari=metin_satirlari,
    )


@app.route("/rapor/karsilastir")
@giris_zorunlu
def rapor_karsilastir():
    siniflar = ogretmen_siniflari(session["ogretmen_id"])
    sid = request.args.get("sinif_id", type=int)
    if siniflar:
        if sid not in [s["id"] for s in siniflar]:
            sid = siniflar[0]["id"]
    else:
        sid = None
    a1 = request.args.get("a1", "").strip()
    a2 = request.args.get("a2", "").strip()
    b1 = request.args.get("b1", "").strip()
    b2 = request.args.get("b2", "").strip()
    sonuc = None
    if sid and a1 and a2 and b1 and b2:
        a1, a2 = _tarih_araligi_duzelt(a1, a2)
        b1, b2 = _tarih_araligi_duzelt(b1, b2)
        sonuc = {
            "donem_a_tik": tik_sayisi_sinif_aralik(sid, a1, a2),
            "donem_a_olumlu": olumlu_sayisi_sinif_aralik(sid, a1, a2),
            "donem_b_tik": tik_sayisi_sinif_aralik(sid, b1, b2),
            "donem_b_olumlu": olumlu_sayisi_sinif_aralik(sid, b1, b2),
            "a_etiket": f"{a1} — {a2}",
            "b_etiket": f"{b1} — {b2}",
        }
    return render_template(
        "rapor_karsilastir.html",
        siniflar=siniflar,
        sinif_id=sid,
        a1=a1,
        a2=a2,
        b1=b1,
        b2=b2,
        sonuc=sonuc,
    )


@app.route("/rapor/sinif/<int:sinif_id>/anonim")
@giris_zorunlu
def rapor_anonim_sinif(sinif_id: int):
    sid_list = [s["id"] for s in ogretmen_siniflari(session["ogretmen_id"])]
    if sinif_id not in sid_list:
        abort(403)
    oz = anonim_sinif_dagilimi(sinif_id)
    sinif_adi = next((s["sinif_adi"] for s in ogretmen_siniflari(session["ogretmen_id"]) if s["id"] == sinif_id), "")
    return render_template(
        "rapor_anonim_sinif.html",
        sinif_id=sinif_id,
        sinif_adi=sinif_adi,
        oz=oz,
    )


@app.route("/ogretmen/denetim")
@giris_zorunlu
def ogretmen_denetim():
    satirlar = denetim_listesi(500)
    return render_template("ogretmen_denetim.html", satirlar=satirlar)


@app.route("/admin/yedek-onay", methods=["POST"])
def admin_yedek_onay():
    if request.form.get("admin_sifre") != ADMIN_SIFRE:
        flash("Yönetici şifresi hatalı.", "error")
        return redirect(url_for("admin_sifreler"))
    admin_meta_set(
        "son_yedek_hatirlatma",
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    )
    flash("Yedek aldığınız kaydedildi; bir sonraki hatırlatma için tarih sıfırlandı.", "success")
    return redirect(url_for("admin_sifreler"))


# ══════════════════════════════════════════════════════════════════════════
# Veri Sifirlama — Admin
# ══════════════════════════════════════════════════════════════════════════

SIFIRLAMA_SIFRESI = "ERENLER2024"

@app.route("/admin/sifirla", methods=["GET"])
@giris_zorunlu
def admin_sifirla_sayfa():
    if not _toplu_sifirlamaya_izinli_mi(session["ogretmen_id"]):
        abort(403)
    return redirect(url_for("analiz_merkezi"))

@app.route("/api/admin/sifirla", methods=["POST"])
@giris_zorunlu
def api_admin_sifirla():
    if not _toplu_sifirlamaya_izinli_mi(session["ogretmen_id"]):
        return jsonify({"ok": False, "sebep": "Yetkisiz"}), 403
    veri = request.json or {}
    girilen = veri.get("sifre", "").strip()
    if girilen != SIFIRLAMA_SIFRESI:
        return jsonify({"ok": False, "sebep": "Sifre yanlis!"}), 403
    sonuc = tum_verileri_sifirla(session["ogretmen_id"], session.get("ogretmen_adi", ""))
    return jsonify(sonuc)


@app.route("/api/admin/yedek/<int:yedek_id>/geri-yukle", methods=["POST"])
@giris_zorunlu
def api_admin_yedek_geri_yukle(yedek_id: int):
    if not _toplu_sifirlamaya_izinli_mi(session["ogretmen_id"]):
        return jsonify({"ok": False, "sebep": "Yetkisiz"}), 403
    veri = request.json or {}
    girilen = veri.get("sifre", "").strip()
    if girilen != SIFIRLAMA_SIFRESI:
        return jsonify({"ok": False, "sebep": "Sifre yanlis!"}), 403
    sonuc = sistem_yedegini_geri_yukle(
        yedek_id,
        session["ogretmen_id"],
        session.get("ogretmen_adi", ""),
    )
    return jsonify(sonuc)


# ══════════════════════════════════════════════════════════════════════════
# Baslama
# ══════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import socket
    port = int(os.environ.get("PORT", 5000))
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]; s.close()
    except Exception:
        ip = "127.0.0.1"

    print(f"\n  http://localhost:{port}  |  http://{ip}:{port}\n")
    app.config["TEMPLATES_AUTO_RELOAD"] = True
    app.run(host="0.0.0.0", port=port, debug=True, threaded=True)

